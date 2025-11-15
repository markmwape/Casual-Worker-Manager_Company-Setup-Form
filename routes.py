from models import WorkerImportLog, ImportField, WorkerCustomFieldValue, ReportField, ActivityLog
from models import Attendance, Task, Worker, Company, User, Workspace, UserWorkspace, MasterAdmin
from flask import render_template, session, redirect, url_for, make_response, abort, request, jsonify, send_file, send_from_directory
from app_init import app, db
from datetime import timedelta
from sqlalchemy import and_
import json
from abilities import upload_file_to_storage, download_file_from_storage
import pandas as pd
import logging
import traceback
import os
from datetime import datetime
from sqlalchemy import and_
import io
import re
from app_init import master_admin_required
from sqlalchemy import func, desc
from subscription_middleware import subscription_required, check_subscription_status, admin_required, feature_required, worker_limit_check
from tier_config import get_tier_spec, get_price_by_product_and_amount, STRIPE_PRICE_MAPPING
import stripe
import hmac
import hashlib
import secrets
import string

# Load secrets before configuring Stripe
try:
    from load_secrets import ensure_secrets_loaded
    ensure_secrets_loaded()
except ImportError as e:
    logging.warning(f"Could not import secret loading functionality: {e}")
except Exception as e:
    logging.warning(f"Error loading secrets: {e}")

# Configure Stripe
stripe.api_key = os.environ.get('STRIPE_SECRET_KEY')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET')
STRIPE_PUBLISHABLE_KEY = os.environ.get('STRIPE_PUBLISHABLE_KEY')

def get_current_company():
    """Helper function to get the current company from workspace session"""
    if 'current_workspace' not in session:
        return None
    
    workspace_id = session['current_workspace']['id']
    return Company.query.filter_by(workspace_id=workspace_id).first()

def get_current_user():
    """Helper function to get the current user from session"""
    if 'user' not in session or 'user_email' not in session['user']:
        return None
    
    user_email = session['user']['user_email']
    return User.query.filter_by(email=user_email).first()

@app.route('/workspace-selection')
def workspace_selection_route():
    # Redirect to the signin page (workspace selection moved there)
    return redirect(url_for('signin_route'))

@app.route('/create-workspace')
def create_workspace_route():
    """Route for creating a new workspace"""
    return render_template('create_workspace.html')

@app.route('/api/user/workspaces', methods=['POST'])
def get_user_workspaces():
    """Get all workspaces associated with a user's email"""
    try:
        data = request.get_json()
        logging.info(f"=== GET USER WORKSPACES CALLED ===")
        logging.info(f"Request data: {data}")
        
        email = data.get('email', '').strip().lower()
        logging.info(f"Normalized email: {email}")
        
        if not email:
            return jsonify({"error": "Email is required"}), 400
        
        # Find user by email
        user = User.query.filter_by(email=email).first()
        logging.info(f"User lookup result: {user.email if user else 'User not found'}")
        
        # If user doesn't exist, create them (they might be signing in for first time)
        if not user:
            user = User(email=email, profile_picture="")
            db.session.add(user)
            db.session.commit()
            logging.info(f"Created new user during workspace lookup: {email} (ID: {user.id})")
        else:
            logging.info(f"Found existing user: {email} (ID: {user.id})")
        
        workspaces_data = []
        processed_workspace_ids = set()
        
        # Get all workspaces this user already has explicit access to
        user_workspaces = UserWorkspace.query.filter_by(user_id=user.id).all()
        logging.info(f"=== CHECKING UserWorkspace TABLE ===")
        logging.info(f"Found {len(user_workspaces)} workspace memberships for user {email} (ID: {user.id})")
        logging.info(f"UserWorkspace records: {[(uw.workspace_id, uw.role) for uw in user_workspaces]}")
        
        for uw in user_workspaces:
            workspace = uw.workspace
            if workspace and workspace.id not in processed_workspace_ids:
                logging.info(f"Adding workspace to list: {workspace.name} (ID: {workspace.id}, Role: {uw.role})")
                workspaces_data.append({
                    "id": workspace.id,
                    "name": workspace.name,
                    "code": workspace.workspace_code,
                    "role": uw.role,
                    "country": workspace.country,
                    "industry": workspace.industry_type,
                    "created_at": workspace.created_at.strftime('%Y-%m-%d') if workspace.created_at else None
                })
                processed_workspace_ids.add(workspace.id)
            elif not workspace:
                logging.warning(f"UserWorkspace {uw.id} has no associated workspace!")
            else:
                logging.warning(f"Workspace {workspace.name} (ID: {workspace.id}) already processed")
        
        # Email is already associated with workspace when sign-in link was sent
        # Just need to fetch the user's workspaces normally
        # (The association happens in /api/workspace/associate-email)
        
        # CROSS-BROWSER FALLBACK: Check for recently created workspaces without real admin
        # This handles sign-in from different browser/device than creation
        # SECURITY: Only show workspaces where THIS user should be admin
        from datetime import datetime, timedelta
        recent_cutoff = datetime.utcnow() - timedelta(hours=1)
        
        recent_workspaces = Workspace.query.filter(
            Workspace.created_at >= recent_cutoff,  # Created within last hour
            Workspace.company_email == email  # SECURITY: Only workspaces created with THIS email
        ).all()
        
        for workspace in recent_workspaces:
            if workspace.id not in processed_workspace_ids:
                # Check if workspace still has placeholder admin for THIS email
                placeholder_user = User.query.get(workspace.created_by)
                if placeholder_user and placeholder_user.email == f"pending_{email}":
                    logging.info(f"Found recently created workspace for {email}: {workspace.name}")
                    workspaces_data.append({
                        "id": workspace.id,
                        "name": workspace.name,
                        "code": workspace.workspace_code,
                        "role": "Admin",  # They will become admin
                        "country": workspace.country,
                        "industry": workspace.industry_type,
                        "created_at": workspace.created_at.strftime('%Y-%m-%d') if workspace.created_at else None,
                        "cross_browser_available": True,  # Flag for cross-browser sign-in
                        "created_minutes_ago": int((datetime.utcnow() - workspace.created_at).total_seconds() / 60)
                    })
                    processed_workspace_ids.add(workspace.id)
        
        # Also check for workspaces where this email should be admin (company_email match)
        pending_admin_workspaces = Workspace.query.filter_by(company_email=email).all()
        for workspace in pending_admin_workspaces:
            if workspace.id not in processed_workspace_ids:
                # Check if this workspace was created with a placeholder user
                placeholder_email = f"pending_{email}"
                placeholder_user = User.query.filter_by(email=placeholder_email).first()
                
                if placeholder_user and workspace.created_by == placeholder_user.id:
                    # This workspace is waiting for this user to become admin
                    workspaces_data.append({
                        "id": workspace.id,
                        "name": workspace.name,
                        "code": workspace.workspace_code,
                        "role": "Admin",  # They will become admin
                        "country": workspace.country,
                        "industry": workspace.industry_type,
                        "created_at": workspace.created_at.strftime('%Y-%m-%d') if workspace.created_at else None,
                        "pending_admin": True  # Flag to indicate this needs admin assignment
                    })
                    processed_workspace_ids.add(workspace.id)
                    logging.info(f"Found pending admin workspace for {email}: {workspace.name}")
        
        return jsonify({
            "success": True,
            "workspaces": workspaces_data,
            "user_email": email
        }), 200
        
    except Exception as e:
        logging.error(f"Error getting user workspaces: {str(e)}")
        logging.error(f"Exception type: {type(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({"error": "Failed to retrieve workspaces"}), 500

@app.route('/api/workspace/associate-email', methods=['POST'])
def associate_workspace_email():
    """Associate an email with a workspace when sign-in link is sent"""
    try:
        data = request.get_json()
        workspace_id = data.get('workspace_id')
        email = data.get('email', '').strip().lower()
        
        logging.info(f"=== ASSOCIATE EMAIL CALLED ===")
        logging.info(f"Request data: {data}")
        logging.info(f"workspace_id: {workspace_id}, email: {email}")
        
        if not workspace_id or not email:
            logging.error(f"Missing required data - workspace_id: {workspace_id}, email: {email}")
            return jsonify({"error": "workspace_id and email are required"}), 400
        
        logging.info(f"Associating email {email} with workspace ID {workspace_id}")
        
        # Get the workspace
        workspace = Workspace.query.get(workspace_id)
        if not workspace:
            logging.error(f"Workspace {workspace_id} not found in database")
            # List all workspaces for debugging
            all_workspaces = Workspace.query.all()
            logging.error(f"All workspaces in DB: {[(w.id, w.name, w.company_email) for w in all_workspaces]}")
            return jsonify({"error": "Workspace not found"}), 404
        
        logging.info(f"Found workspace: {workspace.name}, company_email: {workspace.company_email}, created_by: {workspace.created_by}")
        
        # Get or create the user with this email
        user = User.query.filter_by(email=email).first()
        if not user:
            user = User(email=email, profile_picture="")
            db.session.add(user)
            db.session.flush()
            logging.info(f"Created new user for email association: {email}")
        
        # Check if UserWorkspace relationship already exists
        existing_relationship = UserWorkspace.query.filter_by(
            user_id=user.id,
            workspace_id=workspace_id
        ).first()
        
        if existing_relationship:
            logging.info(f"User {email} already has access to workspace {workspace.name} with role {existing_relationship.role}")
            return jsonify({
                "success": True,
                "message": "User already associated with workspace"
            }), 200
        
        # Check if workspace has a placeholder admin
        placeholder_email = f"pending_{workspace.company_email}"
        logging.info(f"Looking for placeholder user: {placeholder_email}")
        placeholder_user = User.query.filter_by(email=placeholder_email).first()
        
        logging.info(f"Placeholder user found: {placeholder_user.email if placeholder_user else 'None'}")
        logging.info(f"Workspace created_by: {workspace.created_by}, Placeholder user ID: {placeholder_user.id if placeholder_user else 'None'}")
        
        if placeholder_user and workspace.created_by == placeholder_user.id:
            # This workspace is waiting for its admin - transfer ownership
            logging.info(f"✓ TRANSFERRING OWNERSHIP: workspace {workspace.name} from {placeholder_email} to {email}")
            
            workspace.created_by = user.id
            
            # Update company ownership too
            company = Company.query.filter_by(workspace_id=workspace.id).first()
            if company and company.created_by == placeholder_user.id:
                company.created_by = user.id
                logging.info(f"✓ Updated company ownership to user {user.id}")
            
            # Create UserWorkspace with Admin role
            user_workspace = UserWorkspace(
                user_id=user.id,
                workspace_id=workspace.id,
                role='Admin'
            )
            db.session.add(user_workspace)
            logging.info(f"✓ Created UserWorkspace: user_id={user.id}, workspace_id={workspace.id}, role=Admin")
            
            # Delete placeholder user
            db.session.delete(placeholder_user)
            logging.info(f"✓ Deleted placeholder user: {placeholder_email}")
            
            db.session.commit()
            logging.info(f"✓ COMMITTED: All changes saved to database")
            
            # Verify the UserWorkspace was created
            verify_uw = UserWorkspace.query.filter_by(user_id=user.id, workspace_id=workspace.id).first()
            logging.info(f"✓ VERIFICATION: UserWorkspace exists: {verify_uw is not None}, Role: {verify_uw.role if verify_uw else 'N/A'}")
            
            logging.info(f"✓✓✓ Successfully associated email {email} with workspace {workspace.name} as Admin ✓✓✓")
            
            return jsonify({
                "success": True,
                "message": "Email successfully associated as admin",
                "role": "Admin"
            }), 200
        else:
            # Workspace already has a real admin - just add this user as member
            logging.info(f"Adding {email} to workspace {workspace.name} as Member")
            
            user_workspace = UserWorkspace(
                user_id=user.id,
                workspace_id=workspace.id,
                role='Member'
            )
            db.session.add(user_workspace)
            db.session.commit()
            
            logging.info(f"Successfully associated email {email} with workspace {workspace.name} as Member")
            
            return jsonify({
                "success": True,
                "message": "Email successfully associated as member",
                "role": "Member"
            }), 200
            
    except Exception as e:
        logging.error(f"Error associating email with workspace: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({"error": "Failed to associate email with workspace"}), 500

@app.route('/api/send-workspace-email', methods=['POST'])
def send_workspace_email():
    """Send an email with a link to retrieve workspace codes"""
    try:
        data = request.get_json()
        logging.info(f"Send workspace email request data: {data}")
        
        email = data.get('email', '').strip().lower()
        
        if not email:
            return jsonify({"error": "Email is required"}), 400
        
        # Validate email format
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({"error": "Invalid email format"}), 400
        
        # Check if user exists
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "No account found with this email address"}), 404
        
        # Check if user has any workspaces
        user_workspaces = UserWorkspace.query.filter_by(user_id=user.id).all()
        if not user_workspaces:
            return jsonify({"error": "No workspaces found for this email address"}), 404
        
        # Generate a secure token for the email link
        import secrets
        import hashlib
        from datetime import timedelta
        
        # Create a token that expires in 10 minutes
        token_data = f"{email}:{datetime.utcnow().isoformat()}"
        token = hashlib.sha256(token_data.encode()).hexdigest()[:32]
        
        # Store token in session or database (for production, use database)
        # For now, we'll use a simple approach and include email in the link
        
        # Create the link
        base_url = request.host_url.rstrip('/')
        workspace_link = f"{base_url}/signin?email={email}"
        
        # In a real implementation, you would send an actual email here
        # For now, we'll simulate it and return success
        
        subject = "Your Workspace Codes - Embee Accounting"
        html_content = f'''
        <html>
        <body>
            <h2>Retrieve Your Workspace Codes</h2>
            <p>Hello,</p>
            <p>You requested to retrieve your workspace codes. Click the link below to view all workspaces associated with your account:</p>
            <p><a href="{workspace_link}" style="background-color: #1A2B48; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; display: inline-block;">View My Workspaces</a></p>
            <p>This link will expire in 10 minutes for security reasons.</p>
            <p>If you didn't request this, you can safely ignore this email.</p>
            <br>
            <p>Best regards,<br>The Embee Accounting Team</p>
        </body>
        </html>
        '''
        
        # Since Firebase handles authentication emails, we don't need custom email sending
        # Instead, return success and let Firebase handle the sign-in link
        logging.info(f"Firebase will handle authentication email for {email}")
        
        return jsonify({
            "success": True,
            "message": "Sign-in link will be sent via Firebase authentication",
            "email": email
        }), 200
        
    except Exception as e:
        logging.error(f"Error sending workspace email: {str(e)}")
        logging.error(f"Exception type: {type(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": "Failed to send workspace email"}), 500

@app.route('/api/workspace/join', methods=['POST'])
def join_workspace():
    """API endpoint to join a workspace with a code"""
    # Ensure database tables exist before querying
    db.create_all()
    try:
        data = request.get_json()
        logging.info(f"Join workspace request data: {data}")
        
        workspace_code = data.get('workspace_code', '').strip().upper()
        logging.info(f"Looking for workspace with code: {workspace_code}")
        
        if not workspace_code or len(workspace_code) != 16:
            logging.warning(f"Invalid workspace code format: {workspace_code}")
            return jsonify({"error": "Invalid workspace code"}), 400
        
        # Find workspace by code
        workspace = Workspace.query.filter_by(workspace_code=workspace_code).first()
        if not workspace:
            logging.warning(f"Workspace not found for code: {workspace_code}")
            # Log all existing workspace codes for debugging
            all_workspaces = Workspace.query.all()
            logging.info(f"Available workspace codes: {[ws.workspace_code for ws in all_workspaces]}")
            return jsonify({"error": "Workspace not found"}), 404
        
        logging.info(f"Found workspace: {workspace.name} (ID: {workspace.id})")
        
        return jsonify({
            "success": True,
            "workspace": {
                "id": workspace.id,
                "name": workspace.name,
                "code": workspace.workspace_code,
                "address": workspace.address
            }
        }, 200)
        
    except Exception as e:
        logging.error(f"Error joining workspace: {str(e)}")
        logging.error(f"Exception type: {type(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": "Failed to join workspace"}), 500

@app.route('/api/workspace/create', methods=['POST'])
def create_workspace():
    """API endpoint to create a new workspace"""
    try:
        # Log environment variables for debugging
        logging.info("Environment variables:")
        logging.info(f"  INSTANCE_CONNECTION_NAME: {os.environ.get('INSTANCE_CONNECTION_NAME')}")
        logging.info(f"  DB_USER: {os.environ.get('DB_USER', 'cwuser')}")
        logging.info(f"  DB_NAME: {os.environ.get('DB_NAME')}")
        logging.info(f"  K_SERVICE: {os.environ.get('K_SERVICE')}")
        
        # Test database connection with better error handling
        logging.info("Testing database connection before workspace creation")
        try:
            with db.engine.connect() as conn:
                result = conn.execute(db.text("SELECT 1"))
                logging.info(f"Database connection successful: {result.fetchone()}")
        except Exception as db_error:
            logging.error(f"Database connection failed: {str(db_error)}")
            logging.error(f"Database URI: {app.config.get('SQLALCHEMY_DATABASE_URI', 'NOT SET')}")
            # Try to continue anyway - maybe the database will work for actual operations
        
        data = request.get_json()
        logging.info(f"Received workspace creation data: {data}")
        
        company_name = data.get('company_name', '').strip()
        country = data.get('country', '').strip()
        industry_type = data.get('industry_type', '').strip()
        expected_workers = str(data.get('expected_workers', 'not_specified')).strip()
        company_phone = data.get('company_phone', '').strip()
        company_email = data.get('company_email', '').strip()
        
        # Validate all required fields
        required_fields = {
            'company_name': company_name,
            'country': country,
            'industry_type': industry_type,
            'company_phone': company_phone,
            'company_email': company_email
        }
        
        missing_fields = [field for field, value in required_fields.items() if not value]
        if missing_fields:
            return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400
        
        # Validate email format
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, company_email):
            return jsonify({"error": "Invalid email format"}), 400
        
        # Validate expected workers format (if provided)
        valid_worker_ranges = ['below_100', '100_250', '251_500', '501_1000', 'above_1000', 'not_specified']
        if expected_workers not in valid_worker_ranges:
            expected_workers = 'not_specified'
        
        # Create workspace immediately in database with system/placeholder user
        # We'll associate it with the actual user when they sign in
        
        # Create a placeholder user for workspace creation (will be updated during sign-in)
        placeholder_email = f"pending_{company_email}"  # Temporary placeholder
        placeholder_user = User.query.filter_by(email=placeholder_email).first()
        if not placeholder_user:
            placeholder_user = User(email=placeholder_email, profile_picture="")
            db.session.add(placeholder_user)
            db.session.flush()
        
        # Create the workspace immediately
        workspace = Workspace(
            name=company_name,
            country=country,
            industry_type=industry_type,
            expected_workers_string=expected_workers,
            expected_workers=0,
            company_phone=company_phone,
            company_email=company_email,
            address="",
            created_by=placeholder_user.id  # Will be updated to actual user during sign-in
        )
        
        db.session.add(workspace)
        db.session.flush()  # Get the workspace ID
        
        # Generate a unique creation token to link this workspace to whoever signs in next
        import secrets
        import string
        creation_token = ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(32))
        
        # Store the workspace-token mapping in session (temporary)
        # This links the workspace to whoever created it, regardless of their sign-in email
        session['workspace_creation_token'] = creation_token
        session['created_workspace_id'] = workspace.id
        
        # CROSS-BROWSER FALLBACK: Recently created workspaces (within 1 hour) are available
        # This allows sign-in from different browsers/devices
        from datetime import datetime, timedelta
        workspace_expiry = datetime.utcnow() + timedelta(hours=1)
        logging.info(f"Workspace {workspace.name} available for cross-browser sign-in until {workspace_expiry}")
        
        # Create a company for this workspace
        company = Company(
            name=company_name,
            registration_number="",
            address="",
            industry=industry_type,
            phone=company_phone,
            created_by=placeholder_user.id,
            workspace_id=workspace.id
        )
        db.session.add(company)
        db.session.commit()
        
        logging.info(f"Created workspace immediately: {workspace.name} (ID: {workspace.id}) with creation token: {creation_token}")
        
        # Return the workspace data with creation token
        return jsonify({
            "success": True,
            "workspace": {
                "id": workspace.id,
                "name": workspace.name,
                "code": workspace.workspace_code,
                "country": workspace.country,
                "industry_type": workspace.industry_type,
                "company_phone": workspace.company_phone,
                "company_email": workspace.company_email,
                "creation_token": creation_token  # Frontend will store this
            },
            "immediate_creation": True
        }), 200
        
    except Exception as e:
        logging.error(f"Error creating workspace: {str(e)}")
        logging.error(f"Exception type: {type(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({"error": "Failed to create workspace"}), 500

@app.route('/set_session', methods=['POST'])
def set_session():
    try:
        user_data = request.json
        logging.info(f"Setting session with user data: {user_data}")
        
        if not user_data:
            logging.error("No user data received")
            return jsonify({"error": "No user data provided"}), 400
        
        # Extract email from user data
        email = user_data.get('email') or user_data.get('user_email')
        if not email:
            logging.error("No email found in user data")
            return jsonify({"error": "No email provided"}), 400
        
        # Get workspace data if available
        workspace_data = user_data.get('workspace_data')
        
        session['user'] = {
            'user_email': email,
            'display_name': user_data.get('displayName', ''),
            'photo_url': user_data.get('photoURL', ''),
            'uid': user_data.get('uid')
        }

        # Create or update user in database
        with app.app_context():
            user = User.query.filter_by(email=email).first()
            if not user:
                user = User(email=email, profile_picture=user_data.get('photoURL', ''))
                db.session.add(user)
                db.session.commit()
                logging.info(f"Created new user: {email}")
            
            # Handle workspace assignment if workspace data is provided
            if workspace_data:
                logging.info(f"Processing workspace_data: {workspace_data}")
                
                # Check if this is an immediately created workspace that needs admin assignment
                if workspace_data.get('immediate_creation') and workspace_data.get('id'):
                    # Workspace already exists, just need to assign actual admin user
                    workspace = Workspace.query.get(workspace_data['id'])
                    if workspace and workspace.company_email == email:
                        logging.info(f"Found immediately created workspace {workspace.name} for admin {email}")
                        
                        # Update workspace to use actual admin user instead of placeholder
                        placeholder_email = f"pending_{email}"
                        placeholder_user = User.query.filter_by(email=placeholder_email).first()
                        
                        if placeholder_user:
                            # Transfer ownership from placeholder to actual user
                            workspace.created_by = user.id
                            
                            # Update company ownership too
                            company = Company.query.filter_by(workspace_id=workspace.id).first()
                            if company:
                                company.created_by = user.id
                            
                            # Delete placeholder user
                            db.session.delete(placeholder_user)
                        
                        # Add admin user to workspace
                        user_workspace = UserWorkspace(
                            user_id=user.id,
                            workspace_id=workspace.id,
                            role='Admin'
                        )
                        db.session.add(user_workspace)
                        db.session.commit()
                        
                        logging.info(f"Successfully assigned admin {email} to workspace {workspace.name}")
                    else:
                        logging.error(f"Workspace {workspace_data['id']} not found or email mismatch")
                        
                # Check if this is a deferred workspace creation (legacy support)
                elif workspace_data.get('deferred_creation') or workspace_data.get('temp_code'):
                    # Create the workspace now with the actual admin user
                    workspace = Workspace(
                        name=workspace_data.get('company_name') or workspace_data.get('name'),
                        country=workspace_data.get('country'),
                        industry_type=workspace_data.get('industry_type'),
                        expected_workers_string=workspace_data.get('expected_workers_string', 'not_specified'),
                        expected_workers=0,
                        company_phone=workspace_data.get('company_phone'),
                        company_email=workspace_data.get('company_email'),
                        address="",
                        created_by=user.id  # Admin user is the creator
                    )
                    
                    db.session.add(workspace)
                    db.session.flush()  # Get the workspace ID
                    
                    # Create a company for this workspace
                    company = Company(
                        name=workspace_data.get('company_name') or workspace_data.get('name'),
                        registration_number="",
                        address="",
                        industry=workspace_data.get('industry_type'),
                        phone=workspace_data.get('company_phone'),
                        created_by=user.id,
                        workspace_id=workspace.id
                    )
                    db.session.add(company)
                    
                    # Add admin user to workspace
                    user_workspace = UserWorkspace(
                        user_id=user.id,
                        workspace_id=workspace.id,
                        role='Admin'
                    )
                    db.session.add(user_workspace)
                    db.session.commit()
                    
                    # Activity logging removed for now
                    
                    logging.info(f"Successfully created new workspace:")
                    logging.info(f"  - Workspace ID: {workspace.id}")
                    logging.info(f"  - Workspace Name: {workspace.name}")
                    logging.info(f"  - Admin User: {email} (ID: {user.id})")
                    logging.info(f"  - Company: {company.name} (ID: {company.id})")
                    logging.info(f"  - UserWorkspace Role: {user_workspace.role}")
                    
                elif workspace_data.get('id'):
                    # Existing workspace or pending admin assignment
                    workspace = Workspace.query.get(workspace_data['id'])
                    logging.info(f"Queried workspace by ID {workspace_data['id']}: {workspace.name if workspace else 'None'}")
                    logging.info(f"workspace_data flags - session_created: {workspace_data.get('session_created')}, cross_browser_available: {workspace_data.get('cross_browser_available')}, pending_admin: {workspace_data.get('pending_admin')}")
                    
                    if workspace:
                        # Check if this is a session-created workspace (any email can become admin)
                        if workspace_data.get('session_created'):
                            logging.info(f"Processing session-created workspace for {email} to workspace {workspace.name}")
                            
                            # Find and clean up placeholder user
                            placeholder_email = f"pending_{workspace.company_email}"
                            placeholder_user = User.query.filter_by(email=placeholder_email).first()
                            
                            if placeholder_user and workspace.created_by == placeholder_user.id:
                                # Transfer ownership from placeholder to actual user
                                workspace.created_by = user.id
                                
                                # Update company ownership too
                                company = Company.query.filter_by(workspace_id=workspace.id).first()
                                if company and company.created_by == placeholder_user.id:
                                    company.created_by = user.id
                                
                                # Delete placeholder user
                                db.session.delete(placeholder_user)
                                logging.info(f"Cleaned up placeholder user: {placeholder_email}")
                            
                            # Add admin user to workspace
                            user_workspace = UserWorkspace(
                                user_id=user.id,
                                workspace_id=workspace.id,
                                role='Admin'
                            )
                            db.session.add(user_workspace)
                            
                            # Clear the session creation token since it's been used
                            session.pop('workspace_creation_token', None)
                            session.pop('created_workspace_id', None)
                            
                            db.session.commit()
                            logging.info(f"Successfully assigned admin {email} to session-created workspace {workspace.name}")
                            
                        # Check if this is a cross-browser available workspace
                        elif workspace_data.get('cross_browser_available'):
                            logging.info(f"Processing cross-browser workspace assignment for {email} to workspace {workspace.name}")
                            
                            # Find and clean up placeholder user
                            placeholder_user = User.query.get(workspace.created_by)
                            if placeholder_user and placeholder_user.email.startswith('pending_'):
                                # Transfer ownership from placeholder to actual user
                                workspace.created_by = user.id
                                
                                # Update company ownership too
                                company = Company.query.filter_by(workspace_id=workspace.id).first()
                                if company and company.created_by == placeholder_user.id:
                                    company.created_by = user.id
                                
                                # Delete placeholder user
                                db.session.delete(placeholder_user)
                                logging.info(f"Cleaned up placeholder user: {placeholder_user.email}")
                            
                            # Add admin user to workspace
                            user_workspace = UserWorkspace(
                                user_id=user.id,
                                workspace_id=workspace.id,
                                role='Admin'
                            )
                            db.session.add(user_workspace)
                            db.session.commit()
                            
                            logging.info(f"Successfully assigned admin {email} to cross-browser workspace {workspace.name}")
                            
                        # Check if this is a pending admin assignment (email match)
                        elif workspace_data.get('pending_admin'):
                            logging.info(f"Processing pending admin assignment for {email} to workspace {workspace.name}")
                            
                            # Find and clean up placeholder user
                            placeholder_email = f"pending_{email}"
                            placeholder_user = User.query.filter_by(email=placeholder_email).first()
                            
                            if placeholder_user and workspace.created_by == placeholder_user.id:
                                # Transfer ownership from placeholder to actual user
                                workspace.created_by = user.id
                                
                                # Update company ownership too
                                company = Company.query.filter_by(workspace_id=workspace.id).first()
                                if company and company.created_by == placeholder_user.id:
                                    company.created_by = user.id
                                
                                # Delete placeholder user
                                db.session.delete(placeholder_user)
                                logging.info(f"Cleaned up placeholder user: {placeholder_email}")
                            
                            # Add admin user to workspace
                            user_workspace = UserWorkspace(
                                user_id=user.id,
                                workspace_id=workspace.id,
                                role='Admin'
                            )
                            db.session.add(user_workspace)
                            db.session.commit()
                            
                            logging.info(f"Successfully assigned admin {email} to workspace {workspace.name}")
                        else:
                            # Check if user is already in this workspace
                            user_workspace = UserWorkspace.query.filter_by(
                                user_id=user.id, 
                                workspace_id=workspace.id
                            ).first()
                            
                            logging.info(f"Session setting - User ID: {user.id}, Workspace ID: {workspace.id}, Workspace created_by: {workspace.created_by}")
                            logging.info(f"Session setting - Existing UserWorkspace: {user_workspace.role if user_workspace else 'None'}")
                            
                            if not user_workspace:
                                # Check if user is the workspace creator (admin)
                                if workspace.created_by == user.id:
                                    # Workspace creator gets admin access
                                    role = 'Admin'
                                    user_workspace = UserWorkspace(
                                        user_id=user.id,
                                        workspace_id=workspace.id,
                                        role=role
                                    )
                                    db.session.add(user_workspace)
                                    db.session.commit()
                                    logging.info(f"Added workspace creator {email} to workspace {workspace.name} with role {role}")
                                else:
                                    # This should not happen if user was found via /api/user/workspaces
                                    logging.error(f"CRITICAL: User {email} (ID: {user.id}) found via /api/user/workspaces but no UserWorkspace relationship exists for workspace {workspace.name} (ID: {workspace.id})")
                                    
                                    # Let's check if there are any UserWorkspace records for this user
                                    all_user_workspaces = UserWorkspace.query.filter_by(user_id=user.id).all()
                                    logging.error(f"All UserWorkspace records for user {user.id}: {[(uw.workspace_id, uw.role) for uw in all_user_workspaces]}")
                                    
                                    # For now, let's be permissive and create the relationship
                                    user_workspace = UserWorkspace(
                                        user_id=user.id,
                                        workspace_id=workspace.id,
                                        role='Member'
                                    )
                                    db.session.add(user_workspace)
                                    db.session.commit()
                                    logging.info(f"Created missing UserWorkspace relationship for {email} to workspace {workspace.name}")
                            else:
                                logging.info(f"User {email} already has access to workspace {workspace.name} with role {user_workspace.role}")
                        
                        # Ensure a company exists for this workspace
                        existing_company = Company.query.filter_by(workspace_id=workspace.id).first()
                        if not existing_company:
                            # Create a company for this workspace
                            new_company = Company(
                                name=workspace.name,
                                registration_number="",
                                address="",
                                industry=workspace.industry_type or "",
                                phone=workspace.company_phone or "",
                                created_by=user.id,
                                workspace_id=workspace.id
                            )
                            db.session.add(new_company)
                            db.session.commit()
                            logging.info(f"Created company for existing workspace: {workspace.name}")
                    else:
                        logging.error(f"Workspace not found for ID: {workspace_data['id']}")
                else:
                    logging.error("Invalid workspace data provided")
                    return jsonify({"error": "Invalid workspace data"}), 400
            
            # Set workspace info in session if we have a workspace
            logging.info(f"About to set session - workspace in locals: {'workspace' in locals()}")
            if 'workspace' in locals():
                logging.info(f"workspace variable exists, value: {workspace if workspace else 'None'}")
            
            if 'workspace' in locals() and workspace:
                # Always query for the UserWorkspace relationship to get the role
                user_workspace_rel = UserWorkspace.query.filter_by(
                    user_id=user.id,
                    workspace_id=workspace.id
                ).first()
                
                role = user_workspace_rel.role if user_workspace_rel else 'Admin'
                logging.info(f"Setting current_workspace in session: {workspace.name} (ID: {workspace.id}), Role: {role}")
                
                session['current_workspace'] = {
                    'id': workspace.id,
                    'name': workspace.name,
                    'code': workspace.workspace_code,
                    'role': role,
                    'company_email': workspace.company_email,
                    'company_phone': workspace.company_phone
                }
            elif not workspace_data:
                # No workspace data provided - check if user has existing workspaces
                user_workspaces = UserWorkspace.query.filter_by(user_id=user.id).all()
                if user_workspaces:
                    # Auto-select the first workspace
                    first_workspace = user_workspaces[0].workspace
                    if first_workspace:
                        logging.info(f"Auto-selecting existing workspace for session: {first_workspace.name} (ID: {first_workspace.id})")
                        session['current_workspace'] = {
                            'id': first_workspace.id,
                            'name': first_workspace.name,
                            'code': first_workspace.workspace_code,
                            'role': user_workspaces[0].role,
                            'company_email': first_workspace.company_email,
                            'company_phone': first_workspace.company_phone
                        }
                else:
                    logging.info(f"User {email} has no workspaces - session will have no current_workspace")
            else:
                logging.warning(f"No workspace variable found in locals. locals() keys: {list(locals().keys())}")
                if workspace_data:
                    logging.warning(f"workspace_data was provided but no workspace variable was created")

        logging.info(f"Session set successfully: {session['user']}")
        logging.info(f"Session keys: {list(session.keys())}")
        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.error(f"Error setting session: {str(e)}")
        logging.error(f"Exception type: {type(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": "Could not set session"}), 500

@app.route('/api/workspace/payments', methods=['GET'])
def get_workspace_payments():
    """Get workspace subscription and payment information"""
    try:
        if 'current_workspace' not in session:
            return jsonify({'error': 'No active workspace'}), 400
        
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.get(workspace_id)
        
        if not workspace:
            return jsonify({'error': 'Workspace not found'}), 404
        
        # Calculate trial information
        from datetime import date, datetime
        today = date.today()
        is_trial_active = False
        trial_days_left = 0
        
        if workspace.trial_end_date:
            is_trial_active = workspace.trial_end_date > datetime.utcnow()
            if is_trial_active:
                trial_days_left = (workspace.trial_end_date.date() - today).days
        
        # Return workspace payment info
        return jsonify({
            'workspace': {
                'id': workspace.id,
                'name': workspace.name,
                'subscription_status': workspace.subscription_status or 'trial',
                'subscription_tier': workspace.subscription_tier or 'trial',
                'is_trial_active': is_trial_active,
                'trial_days_left': trial_days_left,
                'trial_end_date': workspace.trial_end_date.isoformat() if workspace.trial_end_date else None,
                'subscription_end_date': workspace.subscription_end_date.isoformat() if workspace.subscription_end_date else None
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Error fetching workspace payments: {str(e)}")
        return jsonify({'error': 'Failed to fetch workspace information'}), 500

# Removed: /api/stripe/config - Frontend no longer needs publishable key since using Stripe directly

# Removed: /api/create-checkout-session - Replaced with simpler /api/create-checkout

# Removed: /subscription/success - Stripe handles success pages, webhooks handle subscription updates

@app.route('/stripe/webhook', methods=['POST'])
def stripe_webhook():
    """Handle Stripe webhooks"""
    payload = request.get_data()
    sig_header = request.headers.get('Stripe-Signature')
    
    if not STRIPE_WEBHOOK_SECRET:
        logging.warning("No Stripe webhook secret configured")
        return jsonify({'error': 'Webhook secret not configured'}), 400
    
    try:
        # Verify webhook signature
        event = stripe.Webhook.construct_event(
            payload, sig_header, STRIPE_WEBHOOK_SECRET
        )
        
        logging.info(f"Received Stripe webhook: {event['type']}")
        
        if event['type'] == 'checkout.session.completed':
            handle_checkout_session_completed(event['data']['object'])
        elif event['type'] == 'payment_intent.succeeded':
            handle_payment_intent_succeeded(event['data']['object'])
        elif event['type'] == 'customer.subscription.created':
            handle_subscription_created(event['data']['object'])
        elif event['type'] == 'customer.subscription.updated':
            handle_subscription_updated(event['data']['object'])
        elif event['type'] == 'customer.subscription.deleted':
            handle_subscription_deleted(event['data']['object'])
        elif event['type'] == 'invoice.payment_succeeded':
            handle_payment_succeeded(event['data']['object'])
        elif event['type'] == 'invoice.payment_failed':
            handle_payment_failed(event['data']['object'])
        
        return jsonify({'received': True})
        
    except ValueError as e:
        logging.error(f"Invalid payload in webhook: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Invalid payload'}), 400
    except stripe.error.SignatureVerificationError as e:
        logging.error(f"Invalid signature in webhook: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Invalid signature'}), 400
    except Exception as e:
        logging.error(f"Error processing webhook: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Webhook processing failed'}), 500

def handle_subscription_created(subscription):
    """Handle subscription created webhook"""
    try:
        customer_id = subscription['customer']
        subscription_id = subscription['id']
        
        logging.info(f"Processing subscription created: {subscription_id} for customer {customer_id}")
        
        # Find workspace by customer ID (should already be linked from checkout.session.completed)
        workspace = Workspace.query.filter_by(stripe_customer_id=customer_id).first()
        if not workspace:
            logging.warning(f"No workspace found for customer {customer_id}. This might be normal if checkout.session.completed hasn't processed yet.")
            return None
            
        # Get product and price information
        if subscription.get('items') and subscription['items'].get('data'):
            line_item = subscription['items']['data'][0]
            price_info = line_item.get('price', {})
            product_id = price_info.get('product')
            interval = price_info.get('recurring', {}).get('interval', 'month')
            
            # Map product ID to subscription tier
            product_tier_mapping = {
                'prod_T1ELaKIPUK85by': 'starter',
                'prod_T1EOEHvwiG2NHk': 'growth', 
                'prod_T1EPae2dNy79mG': 'enterprise',
                'prod_T1EQLnddFuPiqg': 'corporate',
            }
            
            subscription_tier = product_tier_mapping.get(product_id, 'starter')
            
            # Determine duration
            duration_days = 365 if interval == 'year' else 30
            
            # Update workspace
            workspace.stripe_subscription_id = subscription_id
            workspace.subscription_status = 'active'
            workspace.subscription_tier = subscription_tier
            workspace.trial_end_date = None  # Clear trial
            
            # Set subscription end date
            from datetime import timedelta
            workspace.subscription_end_date = datetime.utcnow() + timedelta(days=duration_days)
            
            db.session.commit()
            logging.info(f"Subscription created for workspace {workspace.id}: {subscription_tier} tier, {interval}ly billing")
        
    except Exception as e:
        logging.error(f"Error handling subscription created: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()

def handle_subscription_updated(subscription):
    """Handle subscription updated webhook"""
    try:
        subscription_id = subscription['id']
        status = subscription['status']
        
        # Find workspace by subscription ID
        workspace = Workspace.query.filter_by(stripe_subscription_id=subscription_id).first()
        if workspace:
            if status == 'active':
                workspace.subscription_status = 'active'
            elif status == 'canceled':
                workspace.subscription_status = 'canceled'
            elif status in ['past_due', 'unpaid']:
                workspace.subscription_status = 'past_due'
            
            db.session.commit()
            logging.info(f"Subscription updated for workspace {workspace.id}: {status}")
        
    except Exception as e:
        logging.error(f"Error handling subscription updated: {str(e)}")
        db.session.rollback()

def handle_subscription_deleted(subscription):
    """Handle subscription deleted webhook"""
    try:
        subscription_id = subscription['id']
        
        # Find workspace by subscription ID
        workspace = Workspace.query.filter_by(stripe_subscription_id=subscription_id).first()
        if workspace:
            workspace.subscription_status = 'canceled'
            workspace.subscription_end_date = datetime.utcnow()
            
            db.session.commit()
            logging.info(f"Subscription deleted for workspace {workspace.id}")
        
    except Exception as e:
        logging.error(f"Error handling subscription deleted: {str(e)}")
        db.session.rollback()

def handle_payment_succeeded(invoice):
    """Handle successful payment webhook"""
    try:
        subscription_id = invoice.get('subscription')
        if subscription_id:
            workspace = Workspace.query.filter_by(stripe_subscription_id=subscription_id).first()
            if workspace:
                workspace.subscription_status = 'active'
                
                # Extend subscription end date
                from datetime import timedelta
                workspace.subscription_end_date = datetime.utcnow() + timedelta(days=30)
                
                db.session.commit()
                logging.info(f"Payment succeeded for workspace {workspace.id}")
        
    except Exception as e:
        logging.error(f"Error handling payment succeeded: {str(e)}")
        db.session.rollback()

def handle_payment_failed(invoice):
    """Handle failed payment webhook"""
    try:
        subscription_id = invoice.get('subscription')
        if subscription_id:
            workspace = Workspace.query.filter_by(stripe_subscription_id=subscription_id).first()
            if workspace:
                workspace.subscription_status = 'past_due'
                
                db.session.commit()
                logging.info(f"Payment failed for workspace {workspace.id}")
        
    except Exception as e:
        logging.error(f"Error handling payment failed: {str(e)}")
        db.session.rollback()

def handle_checkout_session_completed(session):
    """Handle completed checkout session - this is the most reliable webhook for subscriptions"""
    try:
        customer_id = session.get('customer')
        subscription_id = session.get('subscription')
        
        # Extract workspace code from custom fields
        workspace_code = None
        custom_fields = session.get('custom_fields', [])
        for field in custom_fields:
            field_key = field.get('key', '').lower()
            # Handle various workspace code field naming conventions
            if field_key in ['workspace_code', 'workspacecode', 'workspace-code'] or 'workspace' in field_key:
                workspace_code = field.get('text', {}).get('value') or field.get('dropdown', {}).get('value')
                if workspace_code:
                    break
        
        logging.info(f"Checkout session completed - Customer: {customer_id}, Workspace Code: {workspace_code}")
        
        if not customer_id:
            logging.warning("No customer ID in checkout session")
            return None
        
        # Find workspace by workspace code first, then fallback to customer ID
        workspace = None
        if workspace_code:
            # Look up workspace by the custom workspace code
            workspace = Workspace.query.filter_by(workspace_code=workspace_code).first()
            if not workspace:
                logging.warning(f"No workspace found for workspace code: {workspace_code}")
        
        # Fallback to customer ID lookup if workspace code didn't work
        if not workspace and customer_id:
            workspace = Workspace.query.filter_by(stripe_customer_id=customer_id).first()
            
        if not workspace:
            logging.error(f"CRITICAL: No workspace found for customer {customer_id} or workspace code {workspace_code}")
            
            # Create refund for invalid workspace code
            try:
                if subscription_id:
                    stripe.Subscription.cancel(subscription_id)
                    logging.info(f"Cancelled subscription {subscription_id} for invalid workspace code")
                
                # Create automatic refund for invalid workspace code
                payment_intent_id = session.get('payment_intent')
                if payment_intent_id:
                    refund = stripe.Refund.create(
                        payment_intent=payment_intent_id,
                        reason='requested_by_customer',
                        metadata={
                            'reason': 'Invalid workspace code',
                            'workspace_code': workspace_code,
                            'customer_id': customer_id
                        }
                    )
                    logging.info(f"Created refund {refund.id} for invalid workspace code: {workspace_code}")
                else:
                    logging.warning("No payment_intent found for refund")
                
                # Send alert email to admin
                # send_admin_alert(f"Payment received for invalid workspace code: {workspace_code}")
                
            except Exception as e:
                logging.error(f"Failed to handle invalid workspace code: {str(e)}")
            
            return None
        
        product_id = None
        price_id = None
        subscription_tier = 'starter'  # default
        duration_days = 30  # default monthly
        
        # Get product info from subscription
        if subscription_id:
            try:
                subscription = stripe.Subscription.retrieve(subscription_id)
                if subscription.items.data:
                    line_item = subscription.items.data[0]
                    if line_item.price:
                        product_id = line_item.price.product
                        price_id = line_item.price.id
                        
                        # Determine billing interval
                        interval = line_item.price.recurring.interval if line_item.price.recurring else 'month'
                        duration_days = 365 if interval == 'year' else 30
                        
                        logging.info(f"Checkout completed: Product {product_id}, Price {price_id}")
                        
            except Exception as e:
                logging.error(f"Error retrieving subscription from checkout: {str(e)}")
        
        # Determine tier from product ID
        if product_id:
            subscription_tier = get_subscription_tier_from_product(product_id, price_id)
        else:
            # Fallback: try to get from invoice if subscription lookup failed
            try:
                invoice_id = session.get('invoice')
                if invoice_id:
                    invoice = stripe.Invoice.retrieve(invoice_id)
                    if invoice.lines.data:
                        line_item = invoice.lines.data[0]
                        if line_item.price:
                            product_id = line_item.price.product
                            subscription_tier = get_subscription_tier_from_product(product_id, line_item.price.id)
                            logging.info(f"Got product from invoice fallback: {product_id}")
            except Exception as e:
                logging.warning(f"Invoice fallback failed: {str(e)}")
        
        # Update workspace
        workspace.subscription_status = 'active'
        workspace.subscription_tier = subscription_tier
        workspace.trial_end_date = None  # Clear trial since they paid
        
        # Update Stripe IDs if not already set (important for workspace code lookup)
        if customer_id and not workspace.stripe_customer_id:
            workspace.stripe_customer_id = customer_id
            logging.info(f"Linked customer {customer_id} to workspace {workspace.workspace_code}")
            
        if subscription_id:
            workspace.stripe_subscription_id = subscription_id
        
        # Set subscription end date
        from datetime import timedelta
        workspace.subscription_end_date = datetime.utcnow() + timedelta(days=duration_days)
        
        db.session.commit()
        
        logging.info(f"Checkout session completed: Workspace {workspace.id} upgraded to {subscription_tier} tier (Product: {product_id})")
        
    except Exception as e:
        logging.error(f"Error handling checkout session completed: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()

# Frontend Checkout Endpoints
# =============================================================================
# SIMPLIFIED SUBSCRIPTION MANAGEMENT
# =============================================================================
# All subscription management now handled directly by Stripe billing portal.
# Local checkout/validation endpoints removed.
# Subscription updates come via webhooks only.
# 
# Removed routes:
# - /api/validate-workspace
# - /api/create-checkout  
# - /success
# - /cancel
# =============================================================================

@app.route('/upgrade')
def upgrade_workspace():
    """Direct redirect to Stripe billing portal"""
    stripe_billing_url = "https://billing.stripe.com/p/login/test_5kQaEX5Jg1o5g8lg0SgEg00"
    return redirect(stripe_billing_url)

def handle_payment_intent_succeeded(payment_intent):
    """Handle successful payment intent webhook - upgrade user to paid tier"""
    try:
        # Extract customer ID and payment details
        customer_id = payment_intent.get('customer')
        
        if not customer_id:
            logging.warning("No customer ID in payment intent")
            return None
        
        # Find workspace by customer ID
        workspace = Workspace.query.filter_by(stripe_customer_id=customer_id).first()
        if not workspace:
            logging.warning(f"No workspace found for customer {customer_id}")
            return None
        
        # Get product ID from subscription or invoice
        product_id = None
        price_id = None
        subscription_tier = 'starter'  # default fallback
        
        try:
            # Method 1: Try to get from invoice first
            if payment_intent.get('invoice'):
                invoice = stripe.Invoice.retrieve(payment_intent['invoice'])
                if invoice.lines.data:
                    line_item = invoice.lines.data[0]
                    if line_item.price:
                        product_id = line_item.price.product
                        price_id = line_item.price.id
                        logging.info(f"Got product ID from invoice: {product_id}")
            
            # Method 2: If no product from invoice, try subscription
            if not product_id and payment_intent.get('subscription'):
                subscription = stripe.Subscription.retrieve(payment_intent['subscription'])
                if subscription.items.data:
                    line_item = subscription.items.data[0]
                    if line_item.price:
                        product_id = line_item.price.product
                        price_id = line_item.price.id
                        logging.info(f"Got product ID from subscription: {product_id}")
                        
        except Exception as e:
            logging.warning(f"Could not retrieve product info: {str(e)}")
        
        # Determine subscription tier using helper function
        if product_id:
            amount = payment_intent.get('amount')
            subscription_tier = get_subscription_tier_from_product(product_id, price_id, amount)
            logging.info(f"Product ID {product_id} mapped to tier: {subscription_tier}")
        else:
            logging.warning("No product ID found, using default starter tier")
        
        # Set subscription duration (defaulting to monthly, can be refined based on price data)
        subscription_duration_days = 30  # Monthly default
        
        # Update workspace subscription
        workspace.subscription_status = 'active'
        workspace.subscription_tier = subscription_tier
        
        # Set subscription end date
        from datetime import timedelta
        workspace.subscription_end_date = datetime.utcnow() + timedelta(days=subscription_duration_days)
        
        # Clear trial end date since they're now paid
        workspace.trial_end_date = None
        
        db.session.commit()
        
        logging.info(f"Payment intent succeeded: Workspace {workspace.id} upgraded to {subscription_tier} tier. Product ID: {product_id}")
        
    except Exception as e:
        logging.error(f"Error handling payment intent succeeded: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()

def get_subscription_tier_from_product(product_id, price_id=None, amount=None):
    """Helper function to determine subscription tier from Stripe product/price data"""
    
    # Try to match by product ID and amount first (most accurate)
    if amount and product_id:
        tier = get_price_by_product_and_amount(product_id, amount)
        if tier:
            return tier
    
    # Try to match by price ID from our mapping
    if price_id:
        for price_key, price_info in STRIPE_PRICE_MAPPING.items():
            if price_info['price_id'] == price_id:
                return price_info['tier']
    
    # Fallback to product ID mapping
    product_tier_mapping = {
        'prod_T1ELaKIPUK85by': 'starter',
        'prod_T1EOEHvwiG2NHk': 'growth',
        'prod_T1EPae2dNy79mG': 'enterprise',
        'prod_T1EQLnddFuPiqg': 'corporate',
    }
    
    return product_tier_mapping.get(product_id, 'starter')

# Removed: /test/checkout-session-webhook - Test endpoints not needed in production

# Removed: /test/subscription-webhook - Test endpoints not needed in production

# Removed: /test/payment-intent-webhook and /test_session_route - Test endpoints not needed in production

@app.route('/url')
def url_route():
    """Route to handle URL requests from reports page"""
    return jsonify({'url': request.host_url.rstrip('/')})

@app.route('/runtime-app-settings-url')
def get_runtime_app_settings_url():
    try:
        # Get the base URL from environment or configuration
        base_url = os.environ.get('APP_BASE_URL', 'http://localhost:8080')
        return jsonify({"url": base_url})
    except Exception as e:
        logging.error(f"Error retrieving runtime app settings URL: {str(e)}")
        return jsonify({"error": "Could not retrieve runtime app settings URL"}), 500

@app.route('/signin')
def signin_route():
    return render_template('signin.html')

@app.route('/finishSignin', methods=['GET'])
def finish_signin_route():
    email = request.args.get('email')
    if not email:
        return redirect(url_for('signin_route'))
    return render_template('finishSignin.html')

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static', 'img'), 'logo.png', mimetype='image/png')

@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
def apple_touch_icon():
    return send_from_directory(os.path.join(app.root_path, 'static', 'img'), 'logo.png', mimetype='image/png')

@app.route("/api/company/payout-rate", methods=['POST'])
def update_payout_rate():
    try:
        data = request.get_json()
        new_rate = float(data.get('rate'))
        new_currency = data.get('currency')
        new_symbol = data.get('symbol')
        
        if new_rate <= 0:
            return jsonify({'error': 'Payout rate must be greater than 0'}), 400
            
        if not new_currency or not new_symbol:
            return jsonify({'error': 'Currency and symbol are required'}), 400
            
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
            
        company.daily_payout_rate = new_rate
        company.currency = new_currency
        company.currency_symbol = new_symbol
        db.session.commit()
        
        # Activity logging removed for now
        
        return jsonify({
            'message': 'Payout rate updated successfully',
            'new_rate': new_rate,
            'currency': new_currency,
            'symbol': new_symbol
        }), 200
        
    except ValueError:
        return jsonify({'error': 'Invalid payout rate format'}), 400
    except Exception as e:
        logging.error(f"Error updating payout rate: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update payout rate'}), 500

# Configure logging
logging.basicConfig(level=logging.INFO)

@app.route("/api/activity-logs", methods=['GET'])
def get_activity_logs():
    """Get recent activity logs"""
    try:
        # Check if user is authenticated
        if 'user' not in session or 'user_email' not in session['user']:
            logging.warning("User not authenticated for activity logs")
            return jsonify({'error': 'Not authenticated'}), 401

        # Get query parameters
        limit = request.args.get('limit', 20, type=int)
        limit = min(limit, 100)  # Cap at 100 records

        # Get activities using a simple stub
        activities = []  # Stub implementation - no activities for now
        
        logging.info(f"Returning {len(activities)} activity logs")
        
        return jsonify({
            'activities': activities,
            'total': len(activities)
        })
        
    except Exception as e:
        logging.error(f"Error fetching activity logs: {e}")
        return jsonify({'error': 'Failed to fetch activity logs'}), 500

# Remove the broken activity log filtering code below
# @app.route("/api/activity-logs-old", methods=['GET'])
# Removed broken code that was causing syntax errors

def get_time_ago(date_time):
    """Get human-readable time difference"""
    from datetime import datetime, timezone
    
    now = datetime.utcnow()
    diff = now - date_time
    
    if diff.days > 0:
        return f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
    elif diff.seconds > 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif diff.seconds > 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    else:
        return "Just now"
logger = logging.getLogger(__name__)

@app.teardown_request
def teardown_request(exception=None):
    """Ensure database session is properly closed and rolled back on errors"""
    if exception:
        db.session.rollback()
    db.session.remove()

@app.errorhandler(500)
def internal_server_error(error):
    logger.error(f"500 error: {str(error)}\n{traceback.format_exc()}")
    db.session.rollback()  # Ensure transaction is rolled back
    return render_template('500.html'), 500

@app.route("/")
def landing_route():
    return render_template("landing.html")

@app.route("/terms-of-use")
def terms_of_use_route():
    return render_template("terms_of_use.html")

@app.route("/privacy-policy")
def privacy_policy_route():
    return render_template("privacy_policy.html")

@app.route("/legal-compliance")
def legal_compliance_route():
    return render_template("legal_compliance.html")

@app.route("/logout", methods=['GET'])
def logout_route():
    session.clear()
    return redirect(url_for('landing_route'))

@app.route("/onboarding-test")
def onboarding_test_route():
    """Test route for the onboarding system"""
    return render_template("onboarding_test.html")

@app.route('/api/user/onboarding-status', methods=['GET'])
def get_onboarding_status():
    """Check if user should see onboarding (first-time user detection)"""
    try:
        # Check if user is authenticated
        if 'user' not in session or 'user_email' not in session['user']:
            return jsonify({
                'completed': True,  # Don't show onboarding for unauthenticated users
                'isFirstTime': False
            }), 200
        
        user_email = session['user']['user_email']
        user = User.query.filter_by(email=user_email).first()
        
        if not user:
            return jsonify({
                'completed': True,
                'isFirstTime': False
            }), 200
        
        # Check if user was created recently (within last 5 minutes = likely first session)
        from datetime import datetime, timedelta
        if user.created_at:
            time_since_creation = datetime.utcnow() - user.created_at
            is_new_user = time_since_creation < timedelta(minutes=5)
        else:
            is_new_user = True  # No creation date = assume new
        
        # Check if user has workspaces (new users might not have set up workspaces yet)
        user_workspaces = UserWorkspace.query.filter_by(user_id=user.id).count()
        has_minimal_setup = user_workspaces > 0
        
        # Additional check: see if user has created any workers or tasks (indicates they've used the system)
        current_workspace = session.get('current_workspace')
        has_activity = False
        
        if current_workspace and has_minimal_setup:
            company = Company.query.filter_by(workspace_id=current_workspace['id']).first()
            if company:
                workers_count = Worker.query.filter_by(company_id=company.id).count()
                tasks_count = Task.query.filter_by(company_id=company.id).count()
                has_activity = workers_count > 0 or tasks_count > 0
        
        # Determine if this is a first-time user
        is_first_time = is_new_user and not has_activity
        
        return jsonify({
            'completed': has_activity,  # If they have activity, they've used the system
            'isFirstTime': is_first_time,
            'hasWorkspace': has_minimal_setup,
            'userAge': time_since_creation.total_seconds() if user.created_at else 0
        }), 200
        
    except Exception as e:
        logging.error(f"Error checking onboarding status: {str(e)}")
        # On error, don't show onboarding to be safe
        return jsonify({
            'completed': True,
            'isFirstTime': False
        }), 200

@app.route('/api/onboarding/reset', methods=['POST'])
def reset_onboarding():
    """API endpoint to reset onboarding status for the current user"""
    try:
        # This is primarily a client-side feature using localStorage
        # But we can track it server-side if needed in the future
        return jsonify({
            'success': True,
            'message': 'Onboarding reset successful. Refresh the page to see the tour again.'
        }), 200
    except Exception as e:
        logging.error(f"Error resetting onboarding: {str(e)}")
        return jsonify({'error': 'Failed to reset onboarding'}), 500

@app.route("/signout")
def signout():
    session.clear()
    return redirect(url_for('landing_route'))

@app.route("/api/workers", methods=['GET', 'POST'])
@subscription_required
def handle_workers():
    """Handle workers operations - get list or create new worker"""
    try:
        if request.method == 'GET':
            # Get current company from workspace
            company = get_current_company()
            
            if not company:
                return jsonify({'error': 'Company not found'}), 404
            
            # Get workers for this company with their custom field values
            workers = Worker.query.filter_by(company_id=company.id).all()
            
            # Get custom fields for this company
            custom_fields = ImportField.query.filter_by(company_id=company.id).all()
            
            workers_data = []
            for worker in workers:
                worker_data = {
                    'id': worker.id,
                    'first_name': worker.first_name,
                    'last_name': worker.last_name,
                    'date_of_birth': worker.date_of_birth.isoformat() if worker.date_of_birth else None,
                    'created_at': worker.created_at.isoformat() if worker.created_at else None,
                    'custom_fields': {}
                }
                
                # Add custom field values
                for field in custom_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    worker_data['custom_fields'][field.name] = custom_value.value if custom_value else None
                
                workers_data.append(worker_data)
            
            return jsonify({'workers': workers_data, 'custom_fields': [{'id': f.id, 'name': f.name, 'type': f.field_type} for f in custom_fields]})
        
        elif request.method == 'POST':
            # Create new worker (delegate to existing function)
            return create_worker()
    
    except Exception as e:
        logging.error(f"Error handling workers: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to handle workers'}), 500

@app.route("/api/workers/import", methods=['POST'])
@subscription_required
def import_workers_endpoint():
    """Worker import endpoint - alias for existing import function"""
    return import_workers()

@app.route("/api/reports/verify", methods=['GET'])
@subscription_required
def verify_report_data():
    """Verify if report data exists for given date range"""
    try:
        report_type = request.args.get('type', 'per_day')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        logging.debug(f"Verify report data called: type={report_type}, start={start_date}, end={end_date}")
        
        if not start_date or not end_date:
            logging.warning("Missing dates in verification request")
            return jsonify({'has_data': False, 'message': 'Missing dates'}), 400
        
        # Get current company from workspace
        company = get_current_company()
        if not company:
            logging.warning("No company found for current workspace in verification")
            return jsonify({'has_data': False, 'message': 'Company not found'}), 404
        
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError as e:
            logging.error(f"Invalid date format: {str(e)}")
            return jsonify({'has_data': False, 'message': 'Invalid date format'}), 400
        
        # Check for data based on report type
        if report_type == 'per_day':
            # Look for attendance records marked as Present for per_day tasks
            count = Attendance.query.filter(
                Attendance.company_id == company.id,
                Attendance.date.between(start_date_obj, end_date_obj),
                Attendance.status == 'Present',
                Attendance.task_id.isnot(None)
            ).count()
            logging.debug(f"Per day report verification: found {count} Present records")
        elif report_type == 'per_part':
            # Look for any attendance records with units completed for per_part tasks
            count = Attendance.query.filter(
                Attendance.company_id == company.id,
                Attendance.date.between(start_date_obj, end_date_obj),
                Attendance.units_completed.isnot(None)
            ).count()
            logging.debug(f"Per part report verification: found {count} records with units")
        else:
            logging.warning(f"Invalid report type: {report_type}")
            return jsonify({'has_data': False, 'message': 'Invalid report type'}), 400
        
        has_data = count > 0
        
        logging.info(f"Report verification for {report_type}: {count} records found for company {company.id}")
        
        return jsonify({
            'has_data': has_data,
            'record_count': count,
            'message': f'Found {count} records' if has_data else 'No records found'
        }), 200
        
    except Exception as e:
        logging.error(f"Error verifying report data: {str(e)}")
        logging.error(traceback.format_exc())
        # Return success=true even on error to allow download attempt - backend will catch real errors
        return jsonify({
            'has_data': True,
            'message': 'Unable to verify, but attempting download',
            'error': str(e)
        }), 200

@app.route("/api/reports", methods=['GET'])
@subscription_required
@feature_required('advanced_reporting')
def download_reports():
    """Download reports based on type and date range"""
    try:
        report_type = request.args.get('type', 'per_day')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_type = request.args.get('format', 'csv').lower()
        
        logging.info(f"Report download request: type={report_type}, format={format_type}, dates={start_date} to {end_date}")
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        # Validate format
        if format_type not in ['csv', 'excel', 'xlsx']:
            return jsonify({'error': 'Invalid format. Supported formats: csv, excel, xlsx'}), 400
        
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error("Company not found for current workspace")
            return jsonify({'error': 'Company not found'}), 404
        
        # Parse dates
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError as e:
            logging.error(f"Invalid date format: {str(e)}")
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        # Generate report data based on type
        if report_type == 'per_day':
            logging.info(f"Generating per_day report for company {company.id} from {start_date_obj} to {end_date_obj}")
            report_data = generate_per_day_report(company, start_date_obj, end_date_obj)
        elif report_type == 'per_part':
            logging.info(f"Generating per_part report for company {company.id} from {start_date_obj} to {end_date_obj}")
            report_data = generate_per_part_report(company, start_date_obj, end_date_obj)
        elif report_type == 'per_hour':
            logging.info(f"Generating per_hour report for company {company.id} from {start_date_obj} to {end_date_obj}")
            report_data = generate_per_hour_report(company, start_date_obj, end_date_obj)
        else:
            return jsonify({'error': 'Invalid report type'}), 400
        
        logging.info(f"Report generated with {len(report_data) if report_data else 0} records")
        
        # Check if we have data
        if not report_data:
            logging.warning(f"No data found for {report_type} report from {start_date_obj} to {end_date_obj}")
            return jsonify({'error': 'No data available for the selected date range'}), 400
        
        # Generate file based on format
        if format_type == 'csv':
            response = generate_csv_response(report_data, report_type, start_date_obj, end_date_obj)
        elif format_type in ['excel', 'xlsx']:
            response = generate_excel_response(report_data, report_type, start_date_obj, end_date_obj)
        
        logging.info(f"Report successfully generated and ready for download")
        return response
        
    except Exception as e:
        logging.error(f"Error generating report: {str(e)}")
        logging.error(traceback.format_exc())
        return jsonify({'error': 'Failed to generate report'}), 500

def generate_csv_response(report_data, report_type, start_date, end_date):
    """Generate CSV file response"""
    import io
    import csv
    
    try:
        filename = f'{report_type}_report_{start_date}_to_{end_date}.csv'
        
        output = io.StringIO()
        
        if report_data and len(report_data) > 0:
            # Get fieldnames from first record
            fieldnames = list(report_data[0].keys())
            
            # Create CSV writer with explicit parameters
            writer = csv.DictWriter(output, fieldnames=fieldnames, restval='', extrasaction='ignore')
            
            # Write header
            writer.writeheader()
            
            # Write data rows
            for row in report_data:
                writer.writerow(row)
            
            logging.info(f"CSV generated with {len(report_data)} rows and {len(fieldnames)} columns")
        else:
            logging.warning("No data to write to CSV")
        
        output.seek(0)
        csv_content = output.getvalue()
        
        # Create response with proper encoding
        response = make_response(csv_content.encode('utf-8-sig'))  # utf-8-sig includes BOM for Excel compatibility
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Length'] = len(response.get_data())
        
        logging.info(f"CSV response prepared: {filename} ({len(response.get_data())} bytes)")
        
        return response
        
    except Exception as e:
        logging.error(f"Error generating CSV response: {str(e)}")
        logging.error(traceback.format_exc())
        raise

def generate_excel_response(report_data, report_type, start_date, end_date):
    """Generate Excel file response"""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        filename = f'{report_type}_report_{start_date}_to_{end_date}.xlsx'
        
        # Create DataFrame from report data
        if report_data and len(report_data) > 0:
            df = pd.DataFrame(report_data)
            logging.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
        else:
            logging.warning("No data provided for Excel generation, creating empty workbook")
            df = pd.DataFrame()
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data to 'Report' sheet
            df.to_excel(writer, sheet_name='Report', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Report']
            
            # Define styles
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Format data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Right-align numbers
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Auto-adjust column widths
            for idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(idx)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
            
            logging.info(f"Excel formatting completed: {worksheet.max_row} rows, {worksheet.max_column} columns")
        
        output.seek(0)
        excel_bytes = output.getvalue()
        
        # Create response
        response = make_response(excel_bytes)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Length'] = len(excel_bytes)
        
        logging.info(f"Excel response prepared: {filename} ({len(excel_bytes)} bytes)")
        
        return response
        
    except ImportError as e:
        logging.warning(f"pandas or openpyxl not available, falling back to CSV: {str(e)}")
        return generate_csv_response(report_data, report_type, start_date, end_date)
    except Exception as e:
        logging.error(f"Error generating Excel file: {str(e)}")
        logging.error(traceback.format_exc())
        # Fallback to CSV on error
        try:
            return generate_csv_response(report_data, report_type, start_date, end_date)
        except Exception as csv_error:
            logging.error(f"CSV fallback also failed: {str(csv_error)}")
            return jsonify({'error': 'Failed to generate report file'}), 500

def generate_per_day_report(company, start_date, end_date):
    """Generate per day report data"""
    try:
        logging.info(f"Starting per_day report generation for company {company.id}")
        
        # Get all workers for the company
        workers = Worker.query.filter_by(company_id=company.id).all()
        logging.info(f"Found {len(workers)} workers")
        
        # Get custom fields for this company
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        custom_fields = ReportField.query.filter_by(company_id=company.id).all()
        
        report_data = []
        
        for worker in workers:
            # Find all attendance records for this worker in the date range
            attendance_records = Attendance.query.filter(
                Attendance.worker_id == worker.id,
                Attendance.company_id == company.id,
                Attendance.date.between(start_date, end_date),
                Attendance.status == 'Present'
            ).all()
            
            logging.debug(f"Worker {worker.id}: Found {len(attendance_records)} attendance records")
            
            # Group by task (only per_day tasks)
            per_day_tasks = {}
            for att in attendance_records:
                if att.task and getattr(att.task, 'payment_type', None) == 'per_day':
                    # Only count if task started before or during the end date of the report
                    if att.task.start_date.date() <= end_date:
                        task_id = att.task.id
                        if task_id not in per_day_tasks:
                            per_day_tasks[task_id] = {
                                'task': att.task,
                                'days': 0
                            }
                        per_day_tasks[task_id]['days'] += 1
            
            # Create record for each per_day task
            for task_id, task_data in per_day_tasks.items():
                task = task_data['task']
                days = task_data['days']
                
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', ''),
                    'attendance_days': days,
                    'daily_rate': getattr(task, 'per_day_payout', None) or getattr(company, 'daily_payout_rate', None),
                    'currency': getattr(task, 'per_day_currency', 'USD')
                }
                
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else 'N/A'
                
                # Add custom report fields
                for field in custom_fields:
                    if field.field_type == 'numeric' and field.formula:
                        try:
                            formula = field.formula
                            formula = formula.replace('attendance_days', str(days))
                            formula = formula.replace('daily_rate', str(record['daily_rate'] or 0))
                            record[field.name] = eval(formula)
                        except Exception as e:
                            logging.warning(f"Failed to evaluate formula for field {field.name}: {str(e)}")
                            record[field.name] = 0
                    else:
                        record[field.name] = field.value or ''
                
                report_data.append(record)
        
        logging.info(f"Per_day report generation complete: {len(report_data)} records")
        return report_data
        
    except Exception as e:
        logging.error(f"Error generating per day report: {str(e)}")
        logging.error(traceback.format_exc())
        return []

def generate_per_part_report(company, start_date, end_date):
    """Generate per part report data"""
    try:
        logging.info(f"Starting per_part report generation for company {company.id}")
        
        # Get all workers for the company
        workers = Worker.query.filter_by(company_id=company.id).all()
        logging.info(f"Found {len(workers)} workers")
        
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        custom_fields = ReportField.query.filter_by(company_id=company.id).all()
        
        report_data = []
        
        for worker in workers:
            # Find all attendance records for this worker in the date range for per_part tasks
            attendance_records = Attendance.query.filter(
                Attendance.worker_id == worker.id,
                Attendance.company_id == company.id,
                Attendance.date.between(start_date, end_date)
            ).all()
            
            logging.debug(f"Worker {worker.id}: Found {len(attendance_records)} attendance records")
            
            # Group by task (only per_part tasks)
            per_part_tasks = {}
            for att in attendance_records:
                if att.task and getattr(att.task, 'payment_type', None) == 'per_part':
                    # Only count if task started before or during the end date of the report
                    if att.task.start_date.date() <= end_date:
                        task_id = att.task.id
                        if task_id not in per_part_tasks:
                            per_part_tasks[task_id] = {
                                'task': att.task,
                                'units': 0
                            }
                        # Sum up the units completed
                        units = getattr(att, 'units_completed', 0) or 0
                        per_part_tasks[task_id]['units'] += units
            
            # Create record for each per_part task
            for task_id, task_data in per_part_tasks.items():
                task = task_data['task']
                units = task_data['units']
                
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', ''),
                    'units_completed': units,
                    'per_part_rate': getattr(task, 'per_part_payout', None),
                    'currency': getattr(task, 'per_part_currency', 'USD')
                }
                
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else ''
                
                # Add custom report fields
                for field in custom_fields:
                    if field.field_type == 'numeric' and field.formula:
                        try:
                            formula = field.formula
                            formula = formula.replace('units_completed', str(units))
                            formula = formula.replace('per_part_rate', str(record['per_part_rate'] or 0))
                            record[field.name] = eval(formula)
                        except Exception as e:
                            logging.warning(f"Failed to evaluate formula for field {field.name}: {str(e)}")
                            record[field.name] = 0
                    else:
                        record[field.name] = field.value or ''
                
                report_data.append(record)
        
        logging.info(f"Per_part report generation complete: {len(report_data)} records")
        return report_data
        
    except Exception as e:
        logging.error(f"Error generating per_part report: {str(e)}", exc_info=True)
        return []

def generate_per_hour_report(company, start_date, end_date):
    """Generate per hour report data"""
    try:
        logging.info(f"Starting per_hour report generation for company {company.id}")
        
        # Get all workers for the company
        workers = Worker.query.filter_by(company_id=company.id).all()
        logging.info(f"Found {len(workers)} workers")
        
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        custom_fields = ReportField.query.filter_by(company_id=company.id).all()
        
        report_data = []
        
        for worker in workers:
            # Find all attendance records for this worker in the date range for per_hour tasks
            attendance_records = Attendance.query.filter(
                Attendance.worker_id == worker.id,
                Attendance.company_id == company.id,
                Attendance.date.between(start_date, end_date)
            ).all()
            
            logging.debug(f"Worker {worker.id}: Found {len(attendance_records)} attendance records")
            
            # Group by task (only per_hour tasks)
            per_hour_tasks = {}
            for att in attendance_records:
                if att.task and getattr(att.task, 'payment_type', None) == 'per_hour':
                    # Only count if task started before or during the end date of the report
                    if att.task.start_date.date() <= end_date:
                        task_id = att.task.id
                        if task_id not in per_hour_tasks:
                            per_hour_tasks[task_id] = {
                                'task': att.task,
                                'hours': 0
                            }
                        # Sum up the hours worked
                        hours = getattr(att, 'hours_worked', 0) or 0
                        per_hour_tasks[task_id]['hours'] += hours
            
            # Create record for each per_hour task
            for task_id, task_data in per_hour_tasks.items():
                task = task_data['task']
                hours = task_data['hours']
                
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', ''),
                    'hours_worked': hours,
                    'per_hour_rate': getattr(task, 'per_hour_payout', None),
                    'currency': getattr(task, 'per_hour_currency', 'USD')
                }
                
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else ''
                
                # Add custom report fields
                for field in custom_fields:
                    if field.field_type == 'numeric' and field.formula:
                        try:
                            formula = field.formula
                            formula = formula.replace('hours_worked', str(hours))
                            formula = formula.replace('per_hour_rate', str(record['per_hour_rate'] or 0))
                            record[field.name] = eval(formula)
                        except Exception as e:
                            logging.warning(f"Failed to evaluate formula for field {field.name}: {str(e)}")
                            record[field.name] = 0
                    else:
                        record[field.name] = field.value or ''
                
                report_data.append(record)
        
        logging.info(f"Per_hour report generation complete: {len(report_data)} records")
        return report_data
        
    except Exception as e:
        logging.error(f"Error generating per part report: {str(e)}")
        logging.error(traceback.format_exc())
        return []

@app.route("/api/worker", methods=['POST'])
@subscription_required
@worker_limit_check
def create_worker():
    try:
        data = request.get_json()
        
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Handle date_of_birth
        date_of_birth = None
        if data.get('date_of_birth'):
            try:
                date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except ValueError:
                pass  # Keep date_of_birth as None if invalid format
                
        new_worker = Worker(
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            date_of_birth=date_of_birth,
            company_id=company.id
        )
        db.session.add(new_worker)
        db.session.flush()  # Ensure new_worker.id is available
        # Handle custom fields
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        for field in import_fields:
            # Check for both field.name and custom_field_{field.id} formats
            field_value = data.get(field.name) or data.get(f'custom_field_{field.id}')
            if field_value:
                custom_value = WorkerCustomFieldValue(
                    worker_id=new_worker.id,
                    custom_field_id=field.id,
                    value=field_value
                )
                db.session.add(custom_value)
        
        db.session.commit()
        
        # Activity logging removed for now
        worker_name = f"{new_worker.first_name} {new_worker.last_name}".strip()
        
        return jsonify({'message': 'Worker added successfully'}), 201
        
    except Exception as e:
        logging.error(f"Error creating worker: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to add worker'}), 500
    except Exception as e:
        logging.error(f"Error creating worker: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to add worker'}), 500

@app.route("/api/task", methods=['POST'])
@subscription_required
def create_task():
    try:
        data = request.get_json()
        logging.info(f"Received task creation data: {data}")
        
        # Validate required fields
        if not data.get('name'):
            return jsonify({'error': 'Task name is required'}), 400
        if not data.get('start_date'):
            return jsonify({'error': 'Start date is required'}), 400
        
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error(f"Company not found for current workspace")
            return jsonify({'error': 'Company not found'}), 404
        try:
            # Parse start date
            start_date = datetime.fromisoformat(data['start_date'])
        except ValueError as e:
            logging.error(f"Invalid date format: {data['start_date']}")
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        # Date validation: start date cannot be in the past
        today = datetime.now().date()
        if start_date.date() < today:
            return jsonify({'error': 'Start date cannot be in the past.'}), 400
        # Set status based on date
        status = data.get('status', 'Pending')
        if start_date.date() == today:
            status = 'In Progress'
        elif start_date.date() > today:
            status = 'Pending'
        # Create new task
        new_task = Task(
            name=data['name'],
            description=data.get('description', ''),
            start_date=start_date,
            company_id=company.id,
            status=status,
            payment_type=data.get('payment_type', 'per_day'),
            per_part_payout=data.get('per_part_payout'),
            per_part_currency=data.get('per_part_currency'),
            per_hour_payout=data.get('per_hour_payout'),
            per_hour_currency=data.get('per_hour_currency'),
            per_day_payout=data.get('per_day_payout'),
            per_day_currency=data.get('per_day_currency')
        )
        db.session.add(new_task)
        db.session.commit()
        
        # Activity logging removed for now
        
        logging.info(f"Successfully created task: {new_task.id}")
        
        # Activity logging removed for now
        
        return jsonify({
            'message': 'Task created successfully',
            'task_id': new_task.id
        }), 201
    except Exception as e:
        logging.error(f"Error creating task: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'Failed to create task: {str(e)}'}), 500

@app.route("/api/company", methods=['POST'])
def create_company():
    try:
        data = request.get_json()
        
        # Get current user
        user_email = session['user']['user_email']
        user = User.query.filter_by(email=user_email).first()
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Create new company
        new_company = Company(
            name=data['name'],
            registration_number=data['registration_number'],
            address=data['address'],
            industry=data['industry'],
            phone=data['phone'],
            created_by=user.id
        )
        
        db.session.add(new_company)
        db.session.commit()
        
        return jsonify({'message': 'Company created successfully'}), 201
        
    except Exception as e:
        logging.error(f"Error creating company: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create company'}), 500

@app.route("/api/company/contact", methods=['POST'])
def update_company_contact():
    try:
        # Check if user is logged in
        if 'user' not in session or 'user_email' not in session['user']:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.get_json()
        
        if not data or 'email' not in data or 'phone' not in data:
            return jsonify({'error': 'Email and phone are required'}), 400

        # Get current workspace from the session
        if 'current_workspace' not in session:
            return jsonify({'error': 'No workspace found in session'}), 404
            
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.filter_by(id=workspace_id).first()
        
        if not workspace:
            return jsonify({'error': 'No workspace found for current session'}), 404

        # Update the workspace's email and phone
        workspace.company_email = data['email']
        workspace.company_phone = data['phone']
        
        db.session.commit()
        
        # Update the session data if it exists
        if 'current_workspace' in session and isinstance(session['current_workspace'], dict):
            session['current_workspace']['company_email'] = data['email']
            session['current_workspace']['company_phone'] = data['phone']
            session.modified = True
        
        return jsonify({
            'success': True,
            'message': 'Company contact information updated successfully'
        }), 200
        
    except Exception as e:
        logging.error(f"Error updating company contact: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update company contact information'}), 500

@app.route("/workers", methods=['GET'])
@subscription_required
def workers_route():
    try:
        if 'user' not in session or 'user_email' not in session['user']:
            logger.error("No user in session")
            return redirect(url_for('landing_route'))
            
        user_email = session['user']['user_email']
        logger.info(f"Fetching workers for user: {user_email}")
        
        user = User.query.filter_by(email=user_email).first()
        if not user:
            logger.error(f"User not found: {user_email}")
            return redirect(url_for('landing_route'))
        
        # Get company from current workspace
        if 'current_workspace' not in session:
            logger.error(f"No active workspace for user: {user_email}")
            return redirect(url_for('signin_route'))
            
        workspace_id = session['current_workspace']['id']

        logger.info(f"Looking for company in workspace: {workspace_id}")
        
        # Add error handling for workspace_id being None or invalid
        if not workspace_id:
            logger.error("Workspace ID is None in session")
            session.pop('current_workspace', None)
            return redirect(url_for('signin_route'))
        
        company = Company.query.filter_by(workspace_id=workspace_id).first()
        
        # Default fields to display
        default_fields = [
            {'name': 'First Name', 'type': 'text', 'id': 'first_name'},
            {'name': 'Last Name', 'type': 'text', 'id': 'last_name'},
            {'name': 'Date of Birth', 'type': 'date', 'id': 'date_of_birth'}
        ]
        
        if not company:
            logger.info(f"No company found for workspace: {workspace_id}")
            return render_template('workers.html', workers=[], all_fields=default_fields)

        logger.info(f"Found company: {company.name} (ID: {company.id})")

        try:
            # Get custom fields for this company with error handling
            custom_fields = ImportField.query.filter_by(company_id=company.id).all()
            logger.info(f"Found {len(custom_fields)} custom fields")
            
            # Safe field processing
            all_fields = default_fields.copy()
            for field in custom_fields:
                if field and field.name:  # Ensure field exists and has a name
                    all_fields.append({
                        'name': field.name, 
                        'type': field.field_type or 'text', 
                        'id': field.id
                    })
            
            # Get workers for this company with their custom field values (eager load to avoid N+1 queries)
            from sqlalchemy.orm import joinedload
            workers = Worker.query.filter_by(company_id=company.id).options(
                joinedload(Worker.custom_field_values)
            ).all()
            logger.info(f"Found {len(workers)} workers for company ID: {company.id}")
            
            # Validate worker data to prevent template errors
            valid_workers = []
            for worker in workers:
                try:
                    # Ensure worker has the required attributes
                    if worker:
                        # Test accessing custom_field_values to catch any relationship issues early
                        custom_values_count = len(worker.custom_field_values)
                        logger.info(f"Worker: {worker.first_name or 'Unknown'} {worker.last_name or 'Unknown'}, Custom fields: {custom_values_count}")
                        valid_workers.append(worker)
                except Exception as worker_error:
                    logger.error(f"Error processing worker {worker.id if worker else 'None'}: {str(worker_error)}")
                    # Skip this worker but continue with others
                    continue
            
            logger.info(f"Rendering template with {len(valid_workers)} valid workers and {len(all_fields)} fields")
            return render_template('workers.html', workers=valid_workers, all_fields=all_fields)
            
        except Exception as data_error:
            logger.error(f"Error fetching worker data: {str(data_error)}")
            # Fallback to empty data rather than crashing
            return render_template('workers.html', workers=[], all_fields=default_fields)
            
    except Exception as e:
        logger.error(f"Error in workers_route: {str(e)}\n{traceback.format_exc()}")
        return render_template('500.html'), 500

@app.route("/api/import-field", methods=['GET', 'POST'])
def import_field():
    try:
        # Check if user is authenticated
        if 'user' not in session or 'user_email' not in session['user']:
            logging.error("User not authenticated for import-field endpoint")
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Check if workspace exists in session
        if 'current_workspace' not in session:
            logging.error("No workspace in session for import-field endpoint")
            return jsonify({'error': 'No workspace selected'}), 400
        
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error(f"Company not found for workspace: {session['current_workspace']['id']}")
            return jsonify({'error': 'Company not found'}), 404
            
        if request.method == 'GET':
            fields = ImportField.query.filter_by(company_id=company.id).all()
            return jsonify([{
                'id': field.id,
                'name': field.name,
                'field_type': field.field_type
            } for field in fields])
            
        data = request.get_json()
        
        # Validate required fields
        if not data or 'name' not in data:
            return jsonify({'error': 'Field name is required'}), 400
        
        # Create new import field
        new_field = ImportField(
            company_id=company.id,
            name=data['name'],
            field_type=data.get('type', 'text')  # Default to 'text' if type is not provided
        )
        
        db.session.add(new_field)
        db.session.commit()
        
        return jsonify({
            'id': new_field.id,
            'name': new_field.name,
            'type': new_field.field_type
        }), 201
        
    except Exception as e:
        logging.error(f"Error handling import field: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to handle import field request'}), 500

@app.route("/api/import-field/<int:field_id>", methods=['DELETE'])
def delete_import_field(field_id):
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
            
        # Find and delete field
        field = ImportField.query.filter_by(id=field_id, company_id=company.id).first()
        
        if not field:
            return jsonify({'error': 'Field not found'}), 404
            
        db.session.delete(field)
        db.session.commit()
        
        return jsonify({'message': 'Field deleted successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error deleting import field: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete import field'}), 500

from abilities import llm

@app.route("/api/worker/import", methods=['POST'])
@subscription_required
def import_workers():
    try:
        # Validate upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400

        # Save file using existing helpers – returns a storage identifier we can reuse later
        file_id = upload_file_to_storage(file)
        file_path = file_id  # upload_file_to_storage already returns the full path

        # Analyse the Excel contents – get column names and a small preview (first 5 rows)
        df = pd.read_excel(file_path, na_filter=False).dropna(how='all')
        columns = list(df.columns)
        preview = df.head(5).to_dict(orient='records')

        return jsonify({
            'columns': columns,
            'preview': preview,
            'file_id': file_id
        }), 200
    except Exception as e:
        logging.error(f"Error analysing Excel file: {str(e)}")
        return jsonify({'error': 'Failed to analyse Excel file'}), 500
@app.route("/tasks", methods=['GET'])
@subscription_required
def tasks_route():
    try:
        from datetime import date
        
        # Check if user is authenticated
        if 'user' not in session or 'user_email' not in session['user']:
            logging.error("User not authenticated in tasks_route")
            return redirect(url_for('signin_route'))
        
        # Check if workspace exists
        if 'current_workspace' not in session:
            logging.error("No current workspace in tasks_route")
            return redirect(url_for('signin_route'))
        
        # Get current company from workspace
        company = get_current_company()

        if not company:
            logging.warning("No company found for current workspace in tasks_route")
            return render_template('tasks.html', tasks=[], task_statuses=['Pending', 'In Progress', 'Completed'], today_date=date.today().strftime('%Y-%m-%d'))

        # Get all tasks and sort them: Pending and In Progress first, then Completed
        tasks = Task.query.filter_by(company_id=company.id).all()
        
        # Auto-update task statuses based on start date
        today = date.today()
        for task in tasks:
            # If start date has passed and status is still Pending, change to In Progress
            if task.start_date.date() <= today and task.status == 'Pending':
                task.status = 'In Progress'
        
        # Commit any status changes
        db.session.commit()
        
        # Sort tasks: non-completed tasks first (by created date), then completed tasks (by completion date)
        def sort_key(task):
            if task.status == 'Completed':
                # Completed tasks go to the end, sorted by completion date (newest first)
                return (1, -(task.completion_date.timestamp() if task.completion_date else task.created_at.timestamp()))
            else:
                # Non-completed tasks come first, sorted by creation date (newest first)
                return (0, -task.created_at.timestamp())
        
        tasks = sorted(tasks, key=sort_key)
        
        logging.info(f"Rendering tasks.html with {len(tasks)} tasks for company {company.id}")
        return render_template('tasks.html', tasks=tasks, task_statuses=['Pending', 'In Progress', 'Completed'], today_date=date.today().strftime('%Y-%m-%d'))
    except Exception as e:
        logging.error(f"Error fetching tasks: {str(e)}")
        logging.error(f"Traceback: {traceback.format_exc()}")
        return render_template('500.html'), 500

@app.route("/task/<int:task_id>/attendance", methods=['GET'])
def task_attendance_route(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()

        if not company:
            return render_template('task_attendance.html', task=None, attendance_records=[], workers=[], selected_date=None)

        # Get the specific task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()

        if not task:
            return render_template('500.html'), 500

        # Redirect per_hour tasks to hours worked page
        if task.payment_type == 'per_hour':
            return redirect(url_for('task_hours_worked_route', task_id=task_id))

        # Get all available workers for the company for the dropdown
        available_workers = Worker.query.filter_by(company_id=company.id).all()

        # Get selected date from query parameter, default to today
        from datetime import date
        today = date.today()
        
        # Auto-update task status if start date has passed
        if task.start_date.date() <= today and task.status == 'Pending':
            task.status = 'In Progress'
            db.session.commit()
        
        selected_date_str = request.args.get('date')
        if selected_date_str:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        else:
            selected_date = date.today()

        # If selected date is before task start date, show error and do not show attendance UI
        if selected_date < task.start_date.date():
            return render_template('task_attendance.html', 
                task=task, 
                attendance_records=[],
                workers=[],
                available_workers=available_workers,
                selected_date=selected_date,
                attendance_date_error="Attendance cannot be recorded before the task's start date."
            )

        # Get attendance records for the selected date
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        # Create attendance records for assigned workers if they don't exist
        existing_worker_ids = {record.worker_id for record in attendance_records}
        for worker in task.workers:
            if worker.id not in existing_worker_ids:
                new_attendance = Attendance(
                    worker_id=worker.id,
                    company_id=company.id,
                    date=selected_date,
                    status='Absent',
                    task_id=task.id
                )
                db.session.add(new_attendance)
        db.session.commit()

        # Fetch attendance records again to ensure up-to-date list
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        return render_template('task_attendance.html', 
            task=task, 
            attendance_records=attendance_records,
            workers=task.workers.all(),  # Only show assigned workers
            available_workers=available_workers,  # Pass all available workers for dropdown
            selected_date=selected_date,
            attendance_date_error=None
        )
    except Exception as e:
        logging.error(f"Error fetching task attendance: {str(e)}")
        return render_template('500.html'), 500


@app.route("/home", methods=['GET'])
def home_route():
    try:
        # Check if user is authenticated
        if 'user' not in session or 'user_email' not in session['user']:
            logging.info("User not authenticated, redirecting to signin")
            return redirect(url_for('signin_route'))
        
        # Get current user and workspace
        user_email = session['user']['user_email']
        user = User.query.filter_by(email=user_email).first()
        
        if not user:
            logging.error(f"User not found in database: {user_email}")
            return redirect(url_for('signin_route'))
        
        # Handle case where user has no workspace yet
        if 'current_workspace' not in session:
            logging.info("No current workspace - new user, showing home with workspace creation options")
            # Return home page with empty/default data for new users
            return render_template('home.html', 
                                 company=None, 
                                 total_workers=0, 
                                 total_tasks=0, 
                                 team_members=[],
                                 recent_activities=[],
                                 activity_stats={},
                                 subscription_info={'tier': 'trial', 'status': 'trial'},
                                 usage_stats={'workers_used': 0, 'tasks_used': 0},
                                 subscription_updated=False,
                                 all_fields=[])
        
        # Apply subscription check only after we know user has a workspace
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.get(workspace_id)
        
        if not workspace:
            logging.error(f"Workspace not found: {workspace_id}")
            session.pop('current_workspace', None)
            return redirect(url_for('signin_route'))
        
        # Check subscription status for users with workspaces
        from subscription_middleware import check_subscription_status
        try:
            subscription_valid = check_subscription_status(workspace)
            if not subscription_valid:
                logging.warning(f"Subscription expired for workspace {workspace.name}")
                # Could redirect to upgrade page, but for now continue to show home with limitations
        except Exception as e:
            logging.error(f"Error checking subscription status: {str(e)}")
            # Continue anyway, don't block access due to subscription check errors
        
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.get(workspace_id)
        
        if not workspace:
            logging.error(f"Workspace not found: {workspace_id}")
            session.pop('current_workspace', None)
            return redirect(url_for('signin_route'))
        
        # Verify user has access to this workspace
        user_workspace = UserWorkspace.query.filter_by(
            user_id=user.id,
            workspace_id=workspace_id
        ).first()
        
        if not user_workspace:
            logging.error(f"User {user_email} does not have access to workspace {workspace_id}")
            session.pop('current_workspace', None)
            return redirect(url_for('signin_route'))
        
        company = Company.query.filter_by(workspace_id=workspace_id).first()

        if not company:
            return render_template('home.html', company=None, total_workers=0, total_tasks=0, team_members=[], all_fields=[])

        # Get custom fields for add worker modal
        custom_fields = ImportField.query.filter_by(company_id=company.id).all()
        default_fields = [
            {'name': 'First Name', 'type': 'text', 'id': 'first_name'},
            {'name': 'Last Name', 'type': 'text', 'id': 'last_name'},
            {'name': 'Date of Birth', 'type': 'date', 'id': 'date_of_birth'}
        ]
        all_fields = default_fields + [{'name': field.name, 'type': field.field_type or 'text', 'id': field.id} for field in custom_fields]

        # Calculate total workers for the company
        total_workers = Worker.query.filter_by(company_id=company.id).count()

        # Calculate total tasks for the company
        total_tasks = Task.query.filter_by(company_id=company.id).count()

        # Get team members for this workspace
        user_workspaces = UserWorkspace.query.filter_by(workspace_id=workspace_id).all()
        team_members = []
        for uw in user_workspaces:
            user_obj = User.query.get(uw.user_id)
            if user_obj:
                team_members.append({
                    'id': user_obj.id,
                    'email': user_obj.email,
                    'role': uw.role
                })

        # Get subscription information
        from tier_config import get_tier_spec, format_price
        
        subscription_tier = workspace.subscription_tier or 'trial'
        subscription_status = workspace.subscription_status or 'trial'
        
        subscription_info = {
            'tier': subscription_tier,
            'status': subscription_status,
            'workspace_code': workspace.workspace_code,
            'subscription_end': workspace.subscription_end_date,
            'trial_end': workspace.trial_end_date,
            'is_trial': subscription_status == 'trial' or subscription_tier == 'trial'
        }
        
        # Get tier details
        tier_spec = get_tier_spec(subscription_tier)
        subscription_info.update({
            'tier_name': tier_spec['name'],
            'tier_description': tier_spec['description'],
            'worker_limit': tier_spec.get('worker_limit'),
            'monthly_price': format_price(subscription_tier, 'monthly'),
            'features': tier_spec['features'],
            'support_level': tier_spec.get('support_level', 'Email Support'),
            'upgrade_url': 'https://billing.stripe.com/p/login/test_5kQaEX5Jg1o5g8lg0SgEg00',
            'manage_url': 'https://billing.stripe.com/p/login/test_5kQaEX5Jg1o5g8lg0SgEg00'
        })
        
        # Calculate usage - simplified to focus on workers only
        usage_stats = {
            'workers_used': total_workers,
            'workers_limit': tier_spec.get('worker_limit'),
            'tasks_used': total_tasks,
            'tasks_limit': None  # Removed task limits since they're not the main differentiator
        }

        # Activity logging removed for now
        recent_activities = []
        activity_stats = {}

        # Check if this is a subscription update success
        subscription_updated = request.args.get('subscription_updated', False)
        
        # Prevent repeated confetti - only show once per session
        if subscription_updated and subscription_updated != 'false':
            # Mark as shown in session to prevent repeated displays
            session_key = f"subscription_success_shown_{workspace_id}"
            if not session.get(session_key, False):
                session[session_key] = True
                subscription_updated = True
            else:
                subscription_updated = False
        else:
            subscription_updated = False

        return render_template('home.html', 
                             company=company, 
                             total_workers=total_workers, 
                             total_tasks=total_tasks, 
                             team_members=team_members,
                             recent_activities=recent_activities,
                             activity_stats=activity_stats,
                             subscription_info=subscription_info,
                             usage_stats=usage_stats,
                             subscription_updated=subscription_updated,
                             all_fields=all_fields)
    except Exception as e:
        logging.error(f"Error fetching home data: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return render_template('500.html'), 500

@app.route("/api/team-member", methods=['POST'])
def add_team_member():
    try:
        # Check if current user is an Admin
        if session.get('current_workspace', {}).get('role') != 'Admin':
            return jsonify({'success': False, 'error': 'Only admins can add team members'}), 403
        
        data = request.get_json()
        email = data.get('email', '').strip().lower()  # Normalize email to lowercase
        role = data.get('role')

        if not email or not role:
            return jsonify({'success': False, 'error': 'Email and role are required'}), 400

        # Get current workspace
        workspace_id = session.get('current_workspace', {}).get('id')
        if not workspace_id:
            return jsonify({'success': False, 'error': 'No active workspace'}), 400

        # Check if user already exists
        existing_user = User.query.filter_by(email=email).first()
        
        if existing_user:
            # Check if user is already a member of this workspace
            existing_user_workspace = UserWorkspace.query.filter_by(
                user_id=existing_user.id,
                workspace_id=workspace_id
            ).first()
            
            if existing_user_workspace:
                return jsonify({'success': False, 'error': 'User is already a member of this workspace'}), 400
        else:
            # Create new user if they don't exist
            existing_user = User(email=email)
            db.session.add(existing_user)
            db.session.flush()  # Get the user ID without committing

        # Add user to workspace
        user_workspace = UserWorkspace(
            user_id=existing_user.id,
            workspace_id=workspace_id,
            role=role
        )
        db.session.add(user_workspace)
        db.session.commit()
        
        logging.info(f"Successfully added team member: {email} (User ID: {existing_user.id}) to workspace {workspace_id} with role {role}")

        # Activity logging removed for now

        return jsonify({
            'success': True,
            'message': 'Team member added successfully',
            'user': {
                'id': existing_user.id,
                'email': existing_user.email,
                'role': role
            }
        }), 201

    except Exception as e:
        logging.error(f"Error adding team member: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Failed to add team member: {str(e)}'}), 500

@app.route("/api/team-member/<int:user_id>/role", methods=['PUT'])
def update_team_member_role(user_id):
    try:
        # Check if current user is an Admin
        if session.get('current_workspace', {}).get('role') != 'Admin':
            return jsonify({'error': 'Only admins can update team member roles'}), 403
        
        data = request.get_json()
        new_role = data.get('role')
        
        if not new_role:
            return jsonify({'error': 'Role is required'}), 400
        
        # Get current workspace
        workspace_id = session.get('current_workspace', {}).get('id')
        if not workspace_id:
            return jsonify({'error': 'No active workspace'}), 400
        
        # Find the user's workspace role
        user_workspace = UserWorkspace.query.filter_by(
            user_id=user_id,
            workspace_id=workspace_id
        ).first()
        
        if not user_workspace:
            return jsonify({'error': 'User not found in this workspace'}), 404
        
        user_workspace.role = new_role
        db.session.commit()
        
        return jsonify({
            'message': 'Role updated successfully',
            'user': {
                'id': user_id,
                'email': user_workspace.user.email,
                'role': new_role
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Error updating team member role: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'Failed to update role: {str(e)}'}), 500

@app.route("/api/trial-info", methods=['GET'])
def get_trial_info():
    """Get trial information for current workspace"""
    try:
        if 'current_workspace' not in session:
            return jsonify({'success': False, 'error': 'No active workspace'}), 400
        
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.get(workspace_id)
        
        if not workspace:
            return jsonify({'success': False, 'error': 'Workspace not found'}), 404
        
        # Calculate trial days remaining
        from datetime import date
        today = date.today()
        
        if workspace.trial_end_date:
            days_remaining = (workspace.trial_end_date - today).days
            if days_remaining < 0:
                days_remaining = 0
        else:
            days_remaining = 0
        
        return jsonify({
            'success': True,
            'days_remaining': days_remaining,
            'trial_end_date': workspace.trial_end_date.isoformat() if workspace.trial_end_date else None,
            'is_trial': workspace.subscription_status == 'trial' or workspace.subscription_tier == 'trial'
        })
        
    except Exception as e:
        logging.error(f"Error getting trial info: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to get trial info'}), 500

@app.route("/api/team-members", methods=['GET', 'POST'])
def handle_team_members():
    """Handle team member operations"""
    try:
        if request.method == 'GET':
            # Get team members for current workspace
            workspace_id = session.get('current_workspace', {}).get('id')
            if not workspace_id:
                return jsonify({'success': False, 'error': 'No active workspace'}), 400
            
            user_workspaces = UserWorkspace.query.filter_by(workspace_id=workspace_id).all()
            team_members = []
            for uw in user_workspaces:
                user_obj = User.query.get(uw.user_id)
                if user_obj:
                    team_members.append({
                        'id': user_obj.id,
                        'email': user_obj.email,
                        'role': uw.role
                    })
            
            return jsonify({'success': True, 'team_members': team_members})
        
        elif request.method == 'POST':
            # Check if current user is an Admin
            if session.get('current_workspace', {}).get('role') != 'Admin':
                return jsonify({'success': False, 'error': 'Only admins can add team members'}), 403
            
            data = request.get_json()
            email = data.get('email')
            role = data.get('role')

            if not email or not role:
                return jsonify({'success': False, 'error': 'Email and role are required'}), 400

            # Get current workspace
            workspace_id = session.get('current_workspace', {}).get('id')
            if not workspace_id:
                return jsonify({'success': False, 'error': 'No active workspace'}), 400

            # Check if user already exists
            existing_user = User.query.filter_by(email=email).first()
            
            if existing_user:
                # Check if user is already a member of this workspace
                existing_user_workspace = UserWorkspace.query.filter_by(
                    user_id=existing_user.id,
                    workspace_id=workspace_id
                ).first()
                
                if existing_user_workspace:
                    return jsonify({'success': False, 'error': 'User is already a member of this workspace'}), 400
            else:
                # Create new user if they don't exist
                existing_user = User(email=email)
                db.session.add(existing_user)
                db.session.flush()  # Get the user ID without committing

            # Add user to workspace
            user_workspace = UserWorkspace(
                user_id=existing_user.id,
                workspace_id=workspace_id,
                role=role
            )
            db.session.add(user_workspace)
            db.session.commit()

            # Activity logging removed for now

            return jsonify({
                'success': True,
                'message': 'Team member added successfully',
                'user': {
                    'id': existing_user.id,
                    'email': existing_user.email,
                    'role': role
                }
            }), 201
    
    except Exception as e:
        logging.error(f"Error handling team members: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Failed to handle team members'}), 500

@app.route("/api/team-members/<int:user_id>/role", methods=['PUT'])
def update_team_member_role_api(user_id):
    """Update team member role with proper JSON response"""
    try:
        # Check if current user is an Admin
        if session.get('current_workspace', {}).get('role') != 'Admin':
            return jsonify({'success': False, 'error': 'Only admins can update team member roles'}), 403
        
        # Get JSON data
        if request.content_type == 'application/json':
            data = request.get_json()
            new_role = data.get('role')
        else:
            # Handle form data
            new_role = request.form.get('role')
        
        if not new_role:
            return jsonify({'success': False, 'error': 'Role is required'}), 400
        
        # Get current workspace
        workspace_id = session.get('current_workspace', {}).get('id')
        if not workspace_id:
            return jsonify({'success': False, 'error': 'No active workspace'}), 400
        
        # Find the user's workspace role
        user_workspace = UserWorkspace.query.filter_by(
            user_id=user_id,
            workspace_id=workspace_id
        ).first()
        
        if not user_workspace:
            return jsonify({'success': False, 'error': 'User not found in this workspace'}), 404
        
        user_workspace.role = new_role
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Role updated successfully',
            'user': {
                'id': user_id,
                'email': user_workspace.user.email,
                'role': new_role
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Error updating team member role: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Failed to update role: {str(e)}'}), 500

@app.route("/api/team-members/<int:user_id>", methods=['DELETE'])
def delete_team_member_api(user_id):
    """Delete team member with proper JSON response"""
    try:
        # Check if current user is an Admin
        if session.get('current_workspace', {}).get('role') != 'Admin':
            return jsonify({'success': False, 'error': 'Only admins can delete team members'}), 403
        
        # Get current workspace
        workspace_id = session.get('current_workspace', {}).get('id')
        if not workspace_id:
            return jsonify({'success': False, 'error': 'No active workspace'}), 400
        
        # Get current user
        user_email = session['user']['user_email']
        current_user = User.query.filter_by(email=user_email).first()
        
        # Prevent deleting the current user
        if current_user.id == user_id:
            return jsonify({'success': False, 'error': 'Cannot remove yourself from the workspace'}), 400
        
        # Find the user's workspace membership
        user_workspace = UserWorkspace.query.filter_by(
            user_id=user_id,
            workspace_id=workspace_id
        ).first()
        
        if not user_workspace:
            return jsonify({'success': False, 'error': 'User not found in this workspace'}), 404
            
        # Remove user from workspace (but don't delete the user account)
        db.session.delete(user_workspace)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Team member removed from workspace successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error deleting team member: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Failed to delete team member: {str(e)}'}), 500

@app.route("/api/team-member/<int:user_id>", methods=['DELETE'])
def delete_team_member(user_id):
    try:
        # Check if current user is an Admin
        if session.get('current_workspace', {}).get('role') != 'Admin':
            return jsonify({'error': 'Only admins can delete team members'}), 403
        
        # Get current workspace
        workspace_id = session.get('current_workspace', {}).get('id')
        if not workspace_id:
            return jsonify({'error': 'No active workspace'}), 400
        
        # Get current user
        user_email = session['user']['user_email']
        current_user = User.query.filter_by(email=user_email).first()
        
        # Prevent deleting the current user
        if current_user.id == user_id:
            return jsonify({'error': 'Cannot remove yourself from the workspace'}), 400
        
        # Find the user's workspace membership
        user_workspace = UserWorkspace.query.filter_by(
            user_id=user_id,
            workspace_id=workspace_id
        ).first()
        
        if not user_workspace:
            return jsonify({'error': 'User not found in this workspace'}), 404
            
        # Remove user from workspace (but don't delete the user account)
        db.session.delete(user_workspace)
        db.session.commit()
        
        return jsonify({'message': 'Team member removed from workspace successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error deleting team member: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'Failed to delete team member: {str(e)}'}), 500

@app.route("/attendance", methods=['GET'])
@subscription_required
def attendance_route():
    try:
        from datetime import date, datetime
        # Get current user's company
        user_email = session['user']['user_email']
        company = get_current_company()

        if not company:
            return render_template('attendance.html', attendance_records=[])

        # Get date range from query parameters or use default (last 30 days)
        end_date = datetime.strptime(request.args.get('end_date', date.today().isoformat()), '%Y-%m-%d').date()
        start_date = datetime.strptime(request.args.get('start_date', (end_date - timedelta(days=30)).isoformat()), '%Y-%m-%d').date()

        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date.between(start_date, end_date)
        ).all()

        return render_template('attendance.html', 
            attendance_records=attendance_records, 
            start_date=start_date, 
            end_date=end_date
        )
    except Exception as e:
        logging.error(f"Error fetching attendance: {str(e)}")
        return render_template('500.html'), 500

@app.route("/reports", methods=['GET'])
@subscription_required
@feature_required('advanced_reporting')
def reports_route():
    try:
        from datetime import date, timedelta, datetime
        # Get current company from workspace
        company = get_current_company()

        if not company:
            return render_template('reports.html', report_data={}, custom_fields=[], preview_records=[], import_fields=[], per_day_records=[], per_part_records=[])

        # Get date range from query parameters or use default (last 30 days)
        end_date_param = request.args.get('end_date')
        start_date_param = request.args.get('start_date')
        
        if end_date_param and start_date_param:
            # Use provided dates
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
        else:
            # Use default dates (last 30 days)
            end_date = date.today()
            start_date = end_date - timedelta(days=30)

        # Get custom report fields
        custom_fields = ReportField.query.filter_by(company_id=company.id).all()
        # Get import fields (custom worker fields)
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        # Get all tasks for the company
        tasks = Task.query.filter_by(company_id=company.id).all()
        # Get all workers for the company
        workers = Worker.query.filter_by(company_id=company.id).all()

        # Helper for formula evaluation
        def evaluate_formula(formula, context, custom_fields_dict=None, visited=None):
            if visited is None:
                visited = set()
            try:
                import re
                eval_formula = formula
                field_regex = r'[a-zA-Z_][a-zA-Z0-9_]*'
                field_matches = re.findall(field_regex, eval_formula)
                for field_name in field_matches:
                    if field_name in visited:
                        continue
                    if field_name in context:
                        value = context[field_name]
                    elif custom_fields_dict and field_name in custom_fields_dict:
                        visited.add(field_name)
                        custom_field_formula = custom_fields_dict[field_name]
                        value = evaluate_formula(custom_field_formula, context, custom_fields_dict, visited)
                        visited.remove(field_name)
                    else:
                        value = 0
                    eval_formula = re.sub(r'\b' + re.escape(field_name) + r'\b', str(value), eval_formula)
                result = eval(eval_formula)
                return round(result, 2)
            except Exception as e:
                logging.error(f"Error evaluating formula {formula}: {str(e)}")
                return 0.00

        # Prepare per_day, per_part, and per_hour records
        per_day_records = []
        per_part_records = []
        per_hour_records = []

        for worker in workers:
            # For Per Day: attendance and payout
            attendance_days = 0
            # For Per Part: units completed
            total_units_completed = 0
            # For Per Hour: hours worked
            total_hours_worked = 0
            # Find all attendance records for this worker in the date range
            attendance_records = Attendance.query.filter(
                Attendance.worker_id == worker.id,
                Attendance.company_id == company.id,
                Attendance.date.between(start_date, end_date)
            ).all()
            # Group by task payment type
            per_day_attendance = {}
            per_part_units = {}
            per_hour_hours = {}
            for att in attendance_records:
                if att.task and getattr(att.task, 'payment_type', None) == 'per_day':
                    # Only count attendance if the task is active during the report period
                    # Check if task started before or during the end date of the report
                    if att.task.start_date.date() <= end_date:
                        # Task is relevant to the report period
                        if att.status == 'Present':
                            per_day_attendance.setdefault(att.task_id, 0)
                            per_day_attendance[att.task_id] += 1
                elif att.task and getattr(att.task, 'payment_type', None) == 'per_part':
                    # Only count units if the task is active during the report period
                    if att.task.start_date.date() <= end_date:
                        # Task is relevant to the report period
                        if att.units_completed:
                            per_part_units.setdefault(att.task_id, 0)
                            per_part_units[att.task_id] += att.units_completed
                elif att.task and getattr(att.task, 'payment_type', None) == 'per_hour':
                    # Only count hours if the task is active during the report period
                    if att.task.start_date.date() <= end_date:
                        # Task is relevant to the report period
                        if att.hours_worked:
                            per_hour_hours.setdefault(att.task_id, 0)
                            per_hour_hours[att.task_id] += att.hours_worked
            # For each per_day task, add a record
            for task_id, days in per_day_attendance.items():
                task = Task.query.get(task_id) if task_id else None
                # Compute age
                try:
                    from datetime import date
                    if worker.date_of_birth:
                        today = date.today()
                        age = today.year - worker.date_of_birth.year - ((today.month, today.day) < (worker.date_of_birth.month, worker.date_of_birth.day))
                    else:
                        age = 0
                except Exception:
                    age = 0
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', '') if task else '',
                    'attendance_days': days,
                    'daily_rate': getattr(company, 'daily_payout_rate', None),
                    'age': age,
                }
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else 'N/A'
                # Add custom report fields
                custom_fields_dict = {field.name: field.formula for field in custom_fields if field.field_type == 'numeric'}
                for field in custom_fields:
                    if field.field_type == 'numeric':
                        try:
                            context = {
                                'attendance_days': days,
                                'daily_rate': getattr(company, 'daily_payout_rate', None),
                                **record
                            }
                            result = evaluate_formula(field.formula, context, custom_fields_dict)
                            if field.max_limit is not None:
                                result = min(result, field.max_limit)
                            record[field.name] = result
                        except Exception as e:
                            logging.error(f"Error calculating field {field.name}: {str(e)}")
                            record[field.name] = 0.00
                per_day_records.append(record)
            # For each per_part task, add a record
            for task_id, units in per_part_units.items():
                task = Task.query.get(task_id) if task_id else None
                # Compute age
                try:
                    from datetime import date
                    if worker.date_of_birth:
                        today = date.today()
                        age = today.year - worker.date_of_birth.year - ((today.month, today.day) < (worker.date_of_birth.month, worker.date_of_birth.day))
                    else:
                        age = 0
                except Exception:
                    age = 0
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', '') if task else '',
                    'units_completed': units,
                    'per_part_rate': getattr(task, 'per_part_payout', None) if task else None,
                    'per_part_currency': getattr(task, 'per_part_currency', None) if task else None,
                    'age': age,
                }
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else 'N/A'
                # Add custom report fields for per_part
                per_part_custom_fields = [f for f in custom_fields if f.payout_type in ('per_part', 'both')]
                custom_fields_dict = {field.name: field.formula for field in per_part_custom_fields if field.field_type == 'numeric'}
                for field in per_part_custom_fields:
                    if field.field_type == 'numeric':
                        try:
                            context = {
                                'units_completed': units,
                                'per_part_rate': record['per_part_rate'],
                                **record
                            }
                            result = evaluate_formula(field.formula, context, custom_fields_dict)
                            if field.max_limit is not None:
                                result = min(result, field.max_limit)
                            record[field.name] = result
                        except Exception as e:
                            logging.error(f"Error calculating field {field.name}: {str(e)}")
                            record[field.name] = 0.00
                per_part_records.append(record)
            
            # For each per_hour task, add a record
            for task_id, hours in per_hour_hours.items():
                task = Task.query.get(task_id) if task_id else None
                # Compute age
                try:
                    from datetime import date
                    if worker.date_of_birth:
                        today = date.today()
                        age = today.year - worker.date_of_birth.year - ((today.month, today.day) < (worker.date_of_birth.month, worker.date_of_birth.day))
                    else:
                        age = 0
                except Exception:
                    age = 0
                record = {
                    'first_name': getattr(worker, 'first_name', ''),
                    'last_name': getattr(worker, 'last_name', ''),
                    'task_name': getattr(task, 'name', '') if task else '',
                    'hours_worked': hours,
                    'per_hour_rate': getattr(task, 'per_hour_payout', None) if task else None,
                    'per_hour_currency': getattr(task, 'per_hour_currency', None) if task else None,
                    'age': age,
                }
                # Add import field values
                for field in import_fields:
                    custom_value = WorkerCustomFieldValue.query.filter_by(
                        worker_id=worker.id,
                        custom_field_id=field.id
                    ).first()
                    record[field.name] = custom_value.value if custom_value else 'N/A'
                # Add custom report fields for per_hour
                per_hour_custom_fields = [f for f in custom_fields if f.payout_type in ('per_hour', 'both')]
                custom_fields_dict = {field.name: field.formula for field in per_hour_custom_fields if field.field_type == 'numeric'}
                for field in per_hour_custom_fields:
                    if field.field_type == 'numeric':
                        try:
                            context = {
                                'hours_worked': hours,
                                'per_hour_rate': record['per_hour_rate'],
                                **record
                            }
                            result = evaluate_formula(field.formula, context, custom_fields_dict)
                            if field.max_limit is not None:
                                result = min(result, field.max_limit)
                            record[field.name] = result
                        except Exception as e:
                            logging.error(f"Error calculating field {field.name}: {str(e)}")
                            record[field.name] = 0.00
                per_hour_records.append(record)

        return render_template('reports.html', 
            report_data={}, 
            custom_fields=custom_fields,
            import_fields=import_fields,
            per_day_records=per_day_records,
            per_part_records=per_part_records,
            per_hour_records=per_hour_records,
            start_date=start_date.strftime('%Y-%m-%d'),
            end_date=end_date.strftime('%Y-%m-%d')
        )
    except Exception as e:
        logging.error(f"Error generating reports: {str(e)}")
        db.session.rollback()  # Rollback failed transaction
        return render_template('500.html'), 500
@app.route("/api/worker/analyze-columns", methods=['POST'])
def analyze_columns():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400
        file_id = upload_file_to_storage(file)
        file_path = file_id  # upload_file_to_storage already returns the full path
        df = pd.read_excel(file_path, na_filter=False)
        df = df.dropna(how='all')
        original_columns = list(df.columns)
        logging.info(f"Excel columns found: {original_columns}")
        
        # Return column names and file_id with consistent keys
        return jsonify({
            'columns': original_columns,  # Changed from 'original_columns'
            'file_id': file_id,
            'original_columns': original_columns  # Keep original for backwards compatibility
        }), 200
    except Exception as e:
        logging.error(f"Error analyzing columns: {str(e)}")
        return jsonify({'error': f'Failed to analyze Excel file: {str(e)}'}), 500
@app.route("/api/worker/import-mapped", methods=['POST'])
@subscription_required
def import_mapped_workers():
    file_path = None
    try:
        mapping_str = request.form.get('mapping')
        if not mapping_str:
            return jsonify({'error': 'Mapping not provided'}), 400
        import json
        mapping = json.loads(mapping_str)
        file_id = request.form.get('file_id')
        if not file_id:
            return jsonify({'error': 'File ID not provided'}), 400
        file_path = file_id  # The file_id we stored is the actual path on disk
        
        # Check if file exists
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 400
            
        df = pd.read_excel(file_path, na_filter=True)  # Allow NaN values to be detected
        df = df.dropna(how='all')  # Drop completely empty rows
        
        # Replace NaN values with empty strings for safer text processing
        df = df.fillna('')
        
        # Log the mapping for debugging
        logging.info(f"Column mapping: {mapping}")
        logging.info(f"Processing {len(df)} rows from Excel file")
        
        # Get current user and company
        user_email = session['user']['user_email']
        user = User.query.filter_by(email=user_email).first()
        company = get_current_company()
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Check worker limit before importing
        from tier_config import validate_tier_access, get_worker_limit
        workspace_id = session['current_workspace']['id']
        workspace = Workspace.query.get(workspace_id)
        
        current_worker_count = Worker.query.filter_by(company_id=company.id).count()
        worker_limit = get_worker_limit(workspace.subscription_tier or 'starter')
        
        # Calculate how many workers can be imported
        limit_exceeded = False
        limit_warning = None
        max_can_import = None
        
        if worker_limit is not None:  # None means unlimited
            max_can_import = worker_limit - current_worker_count
            if max_can_import < 0:
                max_can_import = 0
            
            if max_can_import == 0:
                return jsonify({
                    'error': f'Subscription limit reached. Your {workspace.subscription_tier or "starter"} plan allows {worker_limit} workers. You currently have {current_worker_count} workers. To add more workers, please upgrade your subscription plan.'
                }), 403
            
            # Check if we're trying to import more than allowed
            # We'll allow partial imports up to the limit
            if len(df) > max_can_import:
                limit_exceeded = True
                limit_warning = f'You can only import {max_can_import} out of {len(df)} workers due to your subscription limit.'
        
        total_records = len(df)
        successful_imports = 0
        duplicate_records = 0
        error_records = 0
        error_details = []
        limited_by_tier = False
        rows_skipped_due_to_limit = 0
        
        # Process each row individually
        for index, row in df.iterrows():
            try:
                # Check if we've reached the import limit
                if worker_limit is not None and successful_imports >= max_can_import:
                    logging.info(f"Stopped importing at limit. Imported {successful_imports} out of {len(df)} workers.")
                    rows_skipped_due_to_limit = len(df) - index
                    limited_by_tier = True
                    break
                
                # Check if this is an empty row - skip if all mapped values are empty
                row_has_data = False
                for field, excel_col in mapping.items():
                    cell_value = str(row[excel_col]).strip()
                    if cell_value and cell_value.lower() not in ['nan', 'nat', 'none']:
                        row_has_data = True
                        break
                
                if not row_has_data:
                    logging.info(f"Skipping empty row {index + 1}")
                    continue
                
                # Create new worker with mapped data
                worker_data = {}
                custom_field_data = {}
                
                for field, excel_col in mapping.items():
                    cell_value = str(row[excel_col]).strip()
                    
                    # Skip empty or NaN cells
                    if not cell_value or cell_value.lower() in ['nan', 'nat', 'none']:
                        continue
                        
                    # Separate worker fields from custom fields
                    if field in ['first_name', 'last_name', 'date_of_birth']:
                        if field == 'date_of_birth':
                            try:
                                # Parse date value
                                parsed_date = pd.to_datetime(cell_value)
                                if pd.isna(parsed_date):
                                    value = None
                                else:
                                    value = parsed_date.date()
                            except Exception as e:
                                logging.warning(f"Could not parse date_of_birth '{cell_value}': {str(e)}")
                                value = None
                        else:
                            value = cell_value
                        worker_data[field] = value
                    else:
                        custom_field_data[field] = cell_value
                
                logging.debug(f"Processed data for row {index + 1}: {worker_data}")

                # Skip row if required fields are missing
                if not worker_data.get('first_name') and not worker_data.get('last_name'):
                    logging.info(f"Skipping row {index + 1} - no first name or last name provided")
                    continue

                # Duplicate check – skip if a worker with the same first & last name already exists for this company
                if worker_data.get('first_name') and worker_data.get('last_name'):
                    duplicate = Worker.query.filter_by(
                        first_name=worker_data.get('first_name'),
                        last_name=worker_data.get('last_name'),
                        company_id=company.id
                    ).first()
                    if duplicate:
                        duplicate_records += 1
                        logging.info(f"Skipping duplicate worker: {worker_data.get('first_name')} {worker_data.get('last_name')}")
                        continue
                
                # Create a savepoint for this worker
                savepoint = db.session.begin_nested()
                
                try:
                    # Clean and validate all worker data
                    first_name = worker_data.get('first_name', '').strip()
                    last_name = worker_data.get('last_name', '').strip()
                    date_of_birth = worker_data.get('date_of_birth', None)
                    
                    # Ensure no NaT or invalid date values
                    if date_of_birth is not None:
                        if pd.isna(date_of_birth) or str(date_of_birth).lower() in ['nat', 'nan', 'none', '']:
                            date_of_birth = None
                    
                    new_worker = Worker(
                        first_name=first_name,
                        last_name=last_name,
                        date_of_birth=date_of_birth,
                        company_id=company.id,
                        user_id=user.id
                    )
                    db.session.add(new_worker)
                    db.session.flush()
                    
                    # Add custom field values
                    for field_name, value in custom_field_data.items():
                        # If the mapping key is a numeric ID (as the UI sends for existing custom fields), use it directly
                        if str(field_name).isdigit():
                            import_field = ImportField.query.filter_by(
                                id=int(field_name),
                                company_id=company.id
                            ).first()
                        else:
                            import_field = ImportField.query.filter_by(
                                name=field_name,
                                company_id=company.id
                            ).first()
                        
                        if not import_field:
                            import_field = ImportField(
                                name=field_name, 
                                company_id=company.id,
                                field_type='text'  # Default type
                            )
                            db.session.add(import_field)
                            db.session.flush()
                        
                        # Create custom field value
                        custom_value = WorkerCustomFieldValue(
                            worker_id=new_worker.id,
                            custom_field_id=import_field.id,
                            value=value
                        )
                        db.session.add(custom_value)
                    
                    # Commit the savepoint for this worker
                    savepoint.commit()
                    successful_imports += 1
                    logging.info(f"Successfully imported worker: {new_worker.first_name} {new_worker.last_name}")
                    
                except Exception as worker_error:
                    # Rollback this worker only
                    savepoint.rollback()
                    raise worker_error
                    
            except Exception as e:
                # Log the specific error for this row
                error_msg = f"Row {index + 2}: {str(e)}"
                error_records += 1
                error_details.append(error_msg)
                logging.error(f"Error processing row {index + 2}: {str(e)}")
                logging.error(f"Row data: {dict(row)}")
                # Continue with next row
                continue
        
        # Commit the main transaction
        db.session.commit()
        
        # Create import log
        try:
            import_log = WorkerImportLog(
                company_id=company.id,
                filename="Mapped Import",
                total_records=total_records,
                successful_imports=successful_imports,
                duplicate_records=duplicate_records,
                error_records=error_records,
                error_details='\n'.join(error_details) if error_details else None
            )
            db.session.add(import_log)
            db.session.commit()
            logging.info(f"Created import log: {successful_imports} successful, {error_records} errors, {duplicate_records} duplicates")
        except Exception as e:
            logging.error(f"Error creating import log: {str(e)}")
            # Don't fail the entire import if we can't create the log
        
        # Clean up the uploaded file
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logging.info(f"Cleaned up uploaded file: {file_path}")
        except Exception as e:
            logging.warning(f"Could not remove uploaded file {file_path}: {str(e)}")
        
        # Build response with warning if applicable
        response_data = {
            'message': 'Import completed',
            'total_records': total_records,
            'successful_imports': successful_imports,
            'duplicate_records': duplicate_records,
            'error_records': error_records,
            'error_details': error_details
        }
        
        # Add limit warning if applicable
        if limited_by_tier:
            response_data['limit_warning'] = limit_warning
            response_data['limit_exceeded'] = True
            response_data['current_limit'] = worker_limit
            response_data['current_count'] = current_worker_count + successful_imports
            response_data['rows_skipped_due_to_limit'] = rows_skipped_due_to_limit
        
        return jsonify(response_data), 200
        
    except Exception as e:
        logging.error(f"Error importing mapped workers: {str(e)}")
        logging.error(f"Traceback: {traceback.format_exc()}")
        
        # Clean up the uploaded file even on error
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logging.info(f"Cleaned up uploaded file after error: {file_path}")
        except Exception as cleanup_error:
            logging.warning(f"Could not remove uploaded file {file_path} after error: {str(cleanup_error)}")
        
        db.session.rollback()
        return jsonify({'error': f'Failed to import workers: {str(e)}'}), 500

@app.route("/api/worker/<int:worker_id>", methods=['GET', 'DELETE', 'PUT'])
@subscription_required
def handle_single_worker(worker_id):
    """Handle individual worker operations - get, update, or delete"""
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error(f"Company not found for worker {worker_id}")
            return jsonify({'error': 'Company not found'}), 404
        
        # Find the worker
        worker = Worker.query.filter_by(id=worker_id, company_id=company.id).first()
        
        if not worker:
            logging.error(f"Worker {worker_id} not found in company {company.id}")
            return jsonify({'error': 'Worker not found'}), 404
        
        if request.method == 'GET':
            # Get worker data with custom fields
            custom_fields = ImportField.query.filter_by(company_id=company.id).all()
            
            worker_data = {
                'id': worker.id,
                'first_name': worker.first_name,
                'last_name': worker.last_name,
                'date_of_birth': worker.date_of_birth.isoformat() if worker.date_of_birth else None,
                'created_at': worker.created_at.isoformat() if worker.created_at else None,
                'custom_fields': {}
            }
            
            # Add custom field values
            for field in custom_fields:
                custom_value = WorkerCustomFieldValue.query.filter_by(
                    worker_id=worker.id,
                    custom_field_id=field.id
                ).first()
                worker_data['custom_fields'][field.name] = custom_value.value if custom_value else None
            
            return jsonify(worker_data)
        
        elif request.method == 'PUT':
            return update_worker(worker_id)
        
        elif request.method == 'DELETE':
            return delete_worker(worker_id)
    
    except Exception as e:
        logging.error(f"Error handling worker {worker_id}: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()
        return jsonify({'error': 'Failed to handle worker'}), 500

# Helper function for deleting a worker (not a route)
def delete_worker(worker_id):
    """Helper function to delete a worker - called by handle_single_worker route"""
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error(f"Company not found when deleting worker {worker_id}")
            return jsonify({'error': 'Company not found'}), 404
            
        # Find worker and verify they belong to user's company
        worker = Worker.query.filter_by(id=worker_id, company_id=company.id).first()
        
        if not worker:
            logging.error(f"Worker {worker_id} not found for company {company.id}")
            return jsonify({'error': 'Worker not found'}), 404
        
        # Remove worker from all tasks first (task_workers association table)
        worker.tasks = []
        db.session.flush()
            
        # Delete worker (cascade deletes will handle attendance and custom fields)
        db.session.delete(worker)
        db.session.commit()
        
        logging.info(f"Worker {worker_id} deleted successfully from company {company.id}")
        return jsonify({'message': 'Worker deleted successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error deleting worker {worker_id}: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()
        return jsonify({'error': f'Failed to delete worker: {str(e)}'}), 500

# Helper function for updating a worker (not a route)
def update_worker(worker_id):
    """Helper function to update a worker - called by handle_single_worker route"""
    try:
        data = request.get_json()
        # Get current company from workspace
        company = get_current_company()
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        worker = Worker.query.filter_by(id=worker_id, company_id=company.id).first()
        if not worker:
            return jsonify({'error': 'Worker not found'}), 404
        # Handle date_of_birth
        if 'date_of_birth' in data:
            if data['date_of_birth']:
                try:
                    worker.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
                except ValueError:
                    return jsonify({'error': 'Invalid date format for date of birth. Use YYYY-MM-DD'}), 400
            else:
                worker.date_of_birth = None
        
        # Update default fields
        worker.first_name = data.get('first_name', worker.first_name)
        worker.last_name = data.get('last_name', worker.last_name)
        # Update custom fields
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        for field in import_fields:
            if field.name in data:
                custom_value = WorkerCustomFieldValue.query.filter_by(worker_id=worker.id, custom_field_id=field.id).first()
                if custom_value:
                    custom_value.value = data[field.name]
                else:
                    new_value = WorkerCustomFieldValue(worker_id=worker.id, custom_field_id=field.id, value=data[field.name])
                    db.session.add(new_value)
        db.session.commit()
        return jsonify({'message': 'Worker updated successfully'}), 200
    except Exception as e:
        logging.error(f"Error updating worker: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update worker'}), 500

@app.route("/api/task/<int:task_id>/attendance", methods=['POST'])
def update_task_attendance(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            logging.error("Company not found when updating attendance")
            return jsonify({'error': 'Company not found'}), 404
        
        # Get the task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        
        if not task:
            logging.error(f"Task {task_id} not found for company {company.id}")
            return jsonify({'error': 'Task not found'}), 404
        
        data = request.get_json()
        attendance_data = data.get('attendance_data', [])
        selected_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
        
        logging.info(f"Updating attendance for task {task_id} on {selected_date} with {len(attendance_data)} records")
        
        # Validation: Prevent attendance before task start date
        if selected_date < task.start_date.date():
            logging.warning(f"Attempted to record attendance before task start date: {selected_date} < {task.start_date.date()}")
            return jsonify({'error': 'You cannot record attendance before the task start date.'}), 400
        
        updated_count = 0
        created_count = 0
        
        # Update attendance records
        for record in attendance_data:
            worker_id = record.get('worker_id')
            status = record.get('status')
            units_completed = record.get('units_completed')
            hours_worked = record.get('hours_worked')
            
            if not worker_id:
                continue
                
            # Verify worker exists and belongs to company
            worker = Worker.query.filter_by(id=worker_id, company_id=company.id).first()
            if not worker:
                logging.warning(f"Worker {worker_id} not found for company {company.id}")
                continue
            
            # Create new attendance record if it doesn't exist
            attendance = Attendance.query.filter_by(
                worker_id=worker_id, 
                company_id=company.id, 
                date=selected_date,
                task_id=task.id
            ).first()
            
            if attendance:
                # Update existing record
                if status is not None:
                    attendance.status = status
                if units_completed is not None:
                    attendance.units_completed = units_completed
                if hours_worked is not None:
                    attendance.hours_worked = hours_worked
                updated_count += 1
                logging.info(f"Updated attendance for worker {worker_id}: status={status}, units={units_completed}, hours={hours_worked}")
            else:
                # Create new record
                new_attendance = Attendance(
                    worker_id=worker_id,
                    company_id=company.id,
                    date=selected_date,
                    status=status if status is not None else 'Absent',
                    task_id=task.id,
                    units_completed=units_completed,
                    hours_worked=hours_worked
                )
                db.session.add(new_attendance)
                created_count += 1
                logging.info(f"Created new attendance for worker {worker_id}: status={status}, units={units_completed}, hours={hours_worked}")
        
        db.session.commit()
        
        # Verify the records were actually saved
        total_records = Attendance.query.filter_by(
            company_id=company.id,
            date=selected_date,
            task_id=task.id
        ).count()
        
        success_message = f"Attendance updated successfully. Created: {created_count}, Updated: {updated_count}, Total records for this date: {total_records}"
        logging.info(success_message)
        
        return jsonify({
            'message': 'Attendance updated successfully',
            'details': {
                'created': created_count,
                'updated': updated_count,
                'total_records': total_records,
                'date': selected_date.isoformat(),
                'task_name': task.name
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Error updating task attendance: {str(e)}", exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to update task attendance'}), 500

@app.route("/api/task/<int:task_id>/add-worker", methods=['POST'])
def add_worker_to_task(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Get the task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        
        if not task:
            return jsonify({'error': 'Task not found'}), 404
        
        data = request.get_json()
        worker_id = data.get('worker_id')
        
        # Validate worker belongs to the company
        worker = Worker.query.filter_by(id=worker_id, company_id=company.id).first()
        if not worker:
            return jsonify({'error': 'Invalid worker'}), 400
        
        # Prevent duplicate assignment
        if worker in task.workers:
            return jsonify({'error': 'This worker is already assigned to this task.'}), 400
        
        # Add worker to task's workers list
        task.workers.append(worker)
        db.session.commit()
        
        # Create attendance records for all dates from task start date to today
        from datetime import date, timedelta
        current_date = date.today()
        start_date = task.start_date.date()
        
        # Generate all dates from task start to today (inclusive)
        date_range = []
        temp_date = start_date
        while temp_date <= current_date:
            date_range.append(temp_date)
            temp_date += timedelta(days=1)
        
        # Create attendance records for each date (all set to Absent)
        for attendance_date in date_range:
            # Check if attendance record already exists for this date
            existing_attendance = Attendance.query.filter_by(
                worker_id=worker_id,
                company_id=company.id,
                date=attendance_date,
                task_id=task.id
            ).first()
            
            # Only create if it doesn't already exist
            if not existing_attendance:
                attendance = Attendance(
                    worker_id=worker_id,
                    company_id=company.id,
                    date=attendance_date,
                    status='Absent',  # Default status when adding worker
                    task_id=task.id
                )
                db.session.add(attendance)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Worker added to task successfully',
            'worker_id': worker_id
        }), 200
        
    except Exception as e:
        logging.error(f"Error adding worker to task: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to add worker to task'}), 500

@app.route("/api/worker/bulk-delete", methods=['POST'])
@subscription_required
def bulk_delete_workers():
    try:
        data = request.get_json()
        
        if not data:
            logging.error("No JSON data received in bulk delete request")
            return jsonify({'error': 'No data provided'}), 400
            
        # Convert all worker_ids to integers and filter out invalids
        worker_ids = [int(wid) for wid in data.get('worker_ids', []) if str(wid).isdigit()]
        
        if not worker_ids:
            logging.error("No valid worker IDs provided")
            return jsonify({'error': 'No valid worker IDs provided'}), 400

        # Get current company from workspace
        company = get_current_company()
        if not company:
            logging.error("Company not found in bulk delete")
            return jsonify({'error': 'Company not found'}), 404

        logging.info(f"Attempting to bulk delete workers with IDs: {worker_ids} for company {company.id}")

        # Only delete workers belonging to this company
        workers = Worker.query.filter(
            Worker.id.in_(worker_ids), 
            Worker.company_id == company.id
        ).all()
        
        if not workers:
            logging.warning(f"No matching workers found for IDs: {worker_ids}")
            return jsonify({'error': 'No matching workers found'}), 404

        logging.info(f"Found {len(workers)} workers to delete from company {company.id}")

        # Remove workers from all tasks first (task_workers association table)
        for worker in workers:
            worker.tasks = []
        db.session.flush()
        
        # Delete workers (cascade deletes will handle attendance and custom fields)
        for worker in workers:
            db.session.delete(worker)
        
        db.session.commit()
        
        logging.info(f"Successfully deleted {len(workers)} workers")
        return jsonify({'message': f'{len(workers)} worker(s) deleted successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error bulk deleting workers: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()
        return jsonify({'error': f'Failed to bulk delete workers: {str(e)}'}), 500

@app.route("/api/worker/delete-all", methods=['DELETE'])
@subscription_required
def delete_all_workers():
    try:
        # Get current company from workspace
        company = get_current_company()
        if not company:
            logging.error("Company not found when deleting all workers")
            return jsonify({'error': 'Company not found'}), 404

        # Get all workers for this company
        workers = Worker.query.filter_by(company_id=company.id).all()
        
        if not workers:
            logging.info(f"No workers to delete for company {company.id}")
            return jsonify({'message': 'No workers to delete'}), 200

        worker_count = len(workers)
        logging.info(f"Deleting all {worker_count} workers from company {company.id}")

        # Remove all workers from all tasks first (task_workers association table)
        for worker in workers:
            worker.tasks = []
        db.session.flush()
        
        # Delete all workers (cascade deletes will handle attendance and custom fields)
        for worker in workers:
            db.session.delete(worker)
        
        db.session.commit()
        
        logging.info(f"Successfully deleted all {worker_count} workers from company {company.id}")
        return jsonify({'message': f'All {worker_count} worker(s) deleted successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error deleting all workers: {str(e)}")
        logging.error(traceback.format_exc())
        db.session.rollback()
        return jsonify({'error': f'Failed to delete all workers: {str(e)}'}), 500

@app.route("/api/task/<int:task_id>/update-date", methods=['POST'])
def update_task_date(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Get the task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        
        if not task:
            return jsonify({'error': 'Task not found'}), 404
        
        data = request.get_json()
        new_start_date = datetime.fromisoformat(data.get('start_date'))
        
        # Update task start date
        task.start_date = new_start_date
        
        # Update or create attendance records for the new date
        Attendance.query.filter_by(
            company_id=company.id, 
            date=task.start_date.date()
        ).delete()
        
        db.session.commit()
        
        return jsonify({'message': 'Task date updated successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error updating task date: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update task date'}), 500

@app.route("/api/task/<int:task_id>", methods=['DELETE', 'PUT'])
def manage_task(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        # Get the task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        
        if not task:
            return jsonify({'error': 'Task not found'}), 404
        
        if request.method == 'DELETE':
            # Delete attendance records for this task
            Attendance.query.filter_by(
                task_id=task.id
            ).delete()
            
            # Delete task
            db.session.delete(task)
            db.session.commit()
            
            return jsonify({'message': 'Task deleted successfully'}), 200
            
        elif request.method == 'PUT':
            # Update task
            data = request.get_json()
            
            # Validate required fields
            if not data.get('name'):
                return jsonify({'error': 'Task name is required'}), 400
            if not data.get('start_date'):
                return jsonify({'error': 'Start date is required'}), 400
            
            try:
                # Parse start date
                start_date = datetime.fromisoformat(data['start_date'])
            except ValueError as e:
                logging.error(f"Invalid date format: {data['start_date']}")
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
            
            # Check if the original start date has already passed
            today = datetime.now().date()
            original_start_date = task.start_date.date() if task.start_date else None
            
            if original_start_date and original_start_date < today:
                # Original start date has passed - don't allow any date changes
                if start_date.date() != original_start_date:
                    return jsonify({
                        'error': 'Cannot modify start date for tasks that have already started or are in the past. The original start date was ' + original_start_date.strftime('%Y-%m-%d') + '.'
                    }), 400
            else:
                # Original start date is today or in future - validate new date isn't in the past
                if start_date.date() < today:
                    return jsonify({
                        'error': 'Start date cannot be set to a date in the past.'
                    }), 400
            
            # Update task fields
            task.name = data['name']
            task.description = data.get('description', '')
            task.start_date = start_date
            task.payment_type = data.get('payment_type', 'per_day')
            task.per_part_payout = data.get('per_part_payout')
            task.per_part_currency = data.get('per_part_currency')
            task.per_day_payout = data.get('per_day_payout')
            task.per_day_currency = data.get('per_day_currency')
            task.per_hour_payout = data.get('per_hour_payout')
            task.per_hour_currency = data.get('per_hour_currency')
            
            db.session.commit()
            
            return jsonify({'message': 'Task updated successfully'}), 200
        
    except Exception as e:
        logging.error(f"Error managing task: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to manage task'}), 500

@app.route("/api/task/<int:task_id>/status", methods=['POST'])
def update_task_status(task_id):
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'error': 'Status is required'}), 400
            
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'error': 'Task not found'}), 404
            
        # Update status and handle completion date
        task.status = new_status
        if new_status == 'Completed' and not task.completion_date:
            task.completion_date = datetime.utcnow()
        elif new_status != 'Completed':
            task.completion_date = None
            
        db.session.commit()
        
        return jsonify({
            'message': 'Task status updated successfully',
            'status': new_status,
            'completion_date': task.completion_date.strftime('%Y-%m-%d %H:%M:%S') if task.completion_date else None
        }), 200
        
    except Exception as e:
        logging.error(f"Error updating task status: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update task status'}), 500

@app.route("/report/download")
@subscription_required
@feature_required('advanced_reporting')
def download_report():
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
            
        # Get date range from query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Date range required'}), 400
        from datetime import datetime
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Get all workers for the company
        workers = Worker.query.filter_by(company_id=company.id).all()
        import_fields = ImportField.query.filter_by(company_id=company.id).all()
        custom_fields = ReportField.query.filter_by(company_id=company.id).all()

        report_type = request.args.get('report_type')
        # Filter custom fields by report type
        if report_type == 'per_day':
            relevant_custom_fields = [f for f in custom_fields if f.payout_type in ('per_day', 'both')]
        elif report_type == 'per_part':
            relevant_custom_fields = [f for f in custom_fields if f.payout_type in ('per_part', 'both')]
        else:
            relevant_custom_fields = custom_fields  # fallback: include all

        # Helper for formula evaluation
        def evaluate_formula(formula, context, custom_fields_dict=None, visited=None):
            if visited is None:
                visited = set()
            try:
                import re
                eval_formula = formula
                field_regex = r'[a-zA-Z_][a-zA-Z0-9_]*'
                field_matches = re.findall(field_regex, eval_formula)
                for field_name in field_matches:
                    if field_name in visited:
                        continue
                    if field_name in context:
                        value = context[field_name]
                    elif custom_fields_dict and field_name in custom_fields_dict:
                        visited.add(field_name)
                        custom_field_formula = custom_fields_dict[field_name]
                        value = evaluate_formula(custom_field_formula, context, custom_fields_dict, visited)
                        visited.remove(field_name)
                    else:
                        value = 0
                    eval_formula = re.sub(r'\b' + re.escape(field_name) + r'\b', str(value), eval_formula)
                result = eval(eval_formula)
                return round(result, 2)
            except Exception as e:
                logging.error(f"Error evaluating formula {formula}: {str(e)}")
                return 0.00

        # Prepare per_day and per_part records
        per_day_records = []
        per_part_records = []
        for worker in workers:
            attendance_records = Attendance.query.filter(
                Attendance.worker_id == worker.id,
                Attendance.company_id == company.id,
                Attendance.date.between(start_date, end_date)
            ).all()
            per_day_attendance = {}
            per_part_units = {}
            for att in attendance_records:
                if att.task and getattr(att.task, 'payment_type', None) == 'per_day':
                    if att.status == 'Present':
                        per_day_attendance.setdefault(att.task_id, 0)
                        per_day_attendance[att.task_id] += 1
                elif att.task and getattr(att.task, 'payment_type', None) == 'per_part':
                    if att.units_completed:
                        per_part_units.setdefault(att.task_id, 0)
                        per_part_units[att.task_id] += att.units_completed
            # Only add records if there is actual attendance or units completed in the range
            if per_day_attendance:
                for task_id, days in per_day_attendance.items():
                    task = Task.query.get(task_id) if task_id else None
                    # Compute age
                    try:
                        from datetime import date
                        if worker.date_of_birth:
                            today = date.today()
                            age = today.year - worker.date_of_birth.year - ((today.month, today.day) < (worker.date_of_birth.month, worker.date_of_birth.day))
                        else:
                            age = 0
                    except Exception:
                        age = 0
                    row = {
                        'First Name': getattr(worker, 'first_name', ''),
                        'Last Name': getattr(worker, 'last_name', ''),
                        'Task Name': getattr(task, 'name', '') if task else '',
                        'Attendance Days': days,
                        'Daily Rate': getattr(company, 'daily_payout_rate', None),
                        'age': age,
                    }
                    for field in import_fields:
                        custom_value = WorkerCustomFieldValue.query.filter_by(
                            worker_id=worker.id,
                            custom_field_id=field.id
                        ).first()
                        row[field.name] = custom_value.value if custom_value else 'N/A'
                    # Only add relevant custom fields
                    custom_fields_dict = {field.name: field.formula for field in relevant_custom_fields if field.field_type == 'numeric'}
                    for field in relevant_custom_fields:
                        if field.field_type == 'numeric':
                            try:
                                context = {
                                    'attendance_days': days,
                                    'daily_rate': getattr(company, 'daily_payout_rate', None),
                                    **row
                                }
                                result = evaluate_formula(field.formula, context, custom_fields_dict)
                                if field.max_limit is not None:
                                    result = min(result, field.max_limit)
                                row[field.name] = result
                            except Exception as e:
                                logging.error(f"Error calculating field {field.name}: {str(e)}")
                                row[field.name] = 0.00
                    per_day_records.append(row)
            if per_part_units:
                for task_id, units in per_part_units.items():
                    task = Task.query.get(task_id) if task_id else None
                    # Compute age
                    try:
                        from datetime import date
                        if worker.date_of_birth:
                            today = date.today()
                            age = today.year - worker.date_of_birth.year - ((today.month, today.day) < (worker.date_of_birth.month, worker.date_of_birth.day))
                        else:
                            age = 0
                    except Exception:
                        age = 0
                    row = {
                        'First Name': getattr(worker, 'first_name', ''),
                        'Last Name': getattr(worker, 'last_name', ''),
                        'Task Name': getattr(task, 'name', '') if task else '',
                        'Units Completed': units,
                        'Per Part Rate': getattr(task, 'per_part_payout', None) if task else None,
                        'Per Part Currency': getattr(task, 'per_part_currency', None) if task else None,
                        'age': age,
                    }
                    for field in import_fields:
                        custom_value = WorkerCustomFieldValue.query.filter_by(
                            worker_id=worker.id,
                            custom_field_id=field.id
                        ).first()
                        row[field.name] = custom_value.value if custom_value else 'N/A'
                    # Add custom report fields for per_part
                    per_part_custom_fields = [f for f in custom_fields if f.payout_type in ('per_part', 'both')]
                    custom_fields_dict = {field.name: field.formula for field in per_part_custom_fields if field.field_type == 'numeric'}
                    for field in per_part_custom_fields:
                        if field.field_type == 'numeric':
                            try:
                                context = {
                                    'units_completed': units,
                                    'per_part_rate': row['Per Part Rate'],
                                    **row
                                }
                                result = evaluate_formula(field.formula, context, custom_fields_dict)
                                if field.max_limit is not None:
                                    result = min(result, field.max_limit)
                                row[field.name] = result
                            except Exception as e:
                                logging.error(f"Error calculating field {field.name}: {str(e)}")
                                row[field.name] = 0.00
                    per_part_records.append(row)
        # Create Excel file with two sheets
        import pandas as pd
        import io
        output = io.BytesIO()
        # Check for empty report
        if (report_type == 'per_day' and len(per_day_records) == 0) or (report_type == 'per_part' and len(per_part_records) == 0) or (not report_type and len(per_day_records) == 0 and len(per_part_records) == 0):
            return jsonify({'error': "This report is empty. Are you sure you've selected the correct date range?"}), 400
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if report_type == 'per_day':
                df_day = pd.DataFrame(per_day_records)
                # Only include relevant columns (base + import fields + relevant custom fields)
                if per_day_records:
                    base_cols = ['First Name', 'Last Name', 'Task Name', 'Attendance Days', 'Daily Rate']
                    import_cols = [f.name for f in import_fields]
                    custom_cols = [f.name for f in relevant_custom_fields if f.payout_type in ('per_day', 'both')]
                    cols = base_cols + import_cols + custom_cols
                    df_day = df_day.loc[:, [c for c in cols if c in df_day.columns]]
                df_day.to_excel(writer, index=False, sheet_name='Per Day')
            elif report_type == 'per_part':
                df_part = pd.DataFrame(per_part_records)
                if per_part_records:
                    base_cols = ['First Name', 'Last Name', 'Task Name', 'Units Completed', 'Per Part Rate', 'Per Part Currency']
                    import_cols = [f.name for f in import_fields]
                    custom_cols = [f.name for f in relevant_custom_fields if f.payout_type in ('per_part', 'both')]
                    cols = base_cols + import_cols + custom_cols
                    df_part = df_part.loc[:, [c for c in cols if c in df_part.columns]]
                df_part.to_excel(writer, index=False, sheet_name='Per Part')
            else:
                # Both sheets, include all relevant fields
                df_day = pd.DataFrame(per_day_records)
                df_part = pd.DataFrame(per_part_records)
                df_day.to_excel(writer, index=False, sheet_name='Per Day')
                df_part.to_excel(writer, index=False, sheet_name='Per Part')
        output.seek(0)
        from flask import send_file
        from datetime import datetime as dt
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'report_{dt.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logging.error(f"Error generating report: {str(e)}")
        return jsonify({'error': f"Failed to generate report: {str(e)}"}), 500
        # Calculate attendance days for each worker
        worker_attendance = {}
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date.between(start_date, end_date),
            Attendance.status == 'Present'
        ).all()
        
        for record in attendance_records:
            worker_attendance[record.worker_id] = worker_attendance.get(record.worker_id, 0) + 1
        
        # Modify preview records to include attendance days
        # Get custom fields for formula builder
        custom_fields_for_formulas = ReportField.query.filter_by(company_id=company.id).all()
        
        preview_records = []
        recent_attendance = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date.between(start_date, end_date)
        ).order_by(Attendance.date.desc()).limit(5).all()
        
        for record in recent_attendance:
            preview_records.append({
                'worker_name': f"{record.worker.first_name} {record.worker.last_name}",
                'task_name': record.task.name,
                'date': record.date.strftime('%Y-%m-%d'),
                'status': record.status,
                'daily_rate': company.daily_payout_rate,
                'attendance_days': worker_attendance.get(record.worker.id, 0)
            })

@app.route("/api/report-field/<int:field_id>", methods=['GET'])
@subscription_required
@feature_required('advanced_reporting')
def get_report_field(field_id):
    """Get a single report field by ID"""
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        
        field = ReportField.query.filter_by(id=field_id, company_id=company.id).first()
        if not field:
            return jsonify({'error': 'Field not found'}), 404
        
        return jsonify({
            'id': field.id,
            'name': field.name,
            'field_type': field.field_type,
            'formula': field.formula,
            'max_limit': field.max_limit,
            'payout_type': field.payout_type,
            'created_at': field.created_at.isoformat() if field.created_at else None
        }), 200
        
    except Exception as e:
        logging.error(f"Error fetching report field: {str(e)}")
        return jsonify({'error': 'Failed to fetch report field'}), 500

@app.route("/api/report-field", methods=['POST', 'DELETE', 'PUT'])
@subscription_required
@feature_required('advanced_reporting')
def manage_report_field():
    try:
        # Get current company from workspace
        company = get_current_company()
        
        if not company:
            return jsonify({'error': 'Company not found'}), 404
            
        if request.method == 'DELETE':
            field_id = request.args.get('id')
            if not field_id:
                return jsonify({'error': 'Field ID required'}), 400
                
            field = ReportField.query.filter_by(id=field_id, company_id=company.id).first()
            if not field:
                return jsonify({'error': 'Field not found'}), 404
                
            db.session.delete(field)
            db.session.commit()
            return jsonify({'message': 'Field deleted successfully'}), 200
            
        data = request.get_json()
        
        if request.method == 'PUT':
            field_id = request.args.get('id')
            if not field_id:
                return jsonify({'error': 'Field ID required'}), 400
            try:
                field_id = int(field_id)
            except Exception as e:
                logging.error(f"Invalid field_id: {field_id}, error: {e}")
                return jsonify({'error': 'Invalid field ID'}), 400
            field = ReportField.query.filter_by(id=field_id, company_id=company.id).first()
            if not field:
                return jsonify({'error': 'Field not found'}), 404
            # Exclude current field from duplicate name check
            duplicate = ReportField.query.filter(db.func.lower(ReportField.name) == data['name'].lower(), ReportField.company_id == company.id, ReportField.id != field_id).first()
            logging.info(f"PUT /api/report-field: field_id={field_id}, checking for duplicate name={data['name']}, duplicate={duplicate}")
            if duplicate:
                return jsonify({'error': 'A custom field with this name already exists.'}), 400
            field.name = data['name']
            field.formula = data['formula']
            field.max_limit = data.get('max_limit')
            db.session.commit()
            return jsonify({
                'id': field.id,
                'name': field.name,
                'field_type': 'numeric',
                'formula': field.formula,
                'max_limit': field.max_limit
            }), 200
        
        # POST method
        logging.info(f"Creating new report field: name={data.get('name')}, formula={data.get('formula')}, payout_type={data.get('payout_type')}")
        
        # Check for duplicate field name (case-insensitive)
        existing = ReportField.query.filter(db.func.lower(ReportField.name) == data['name'].lower(), ReportField.company_id == company.id).first()
        if existing:
            logging.warning(f"Duplicate field name detected: {data['name']}")
            return jsonify({'error': 'A custom field with this name already exists.'}), 400
        
        new_field = ReportField(
            company_id=company.id,
            name=data['name'],
            field_type='numeric',
            formula=data['formula'],
            max_limit=data.get('max_limit'),
            payout_type=data.get('payout_type', 'per_day')
        )
        
        db.session.add(new_field)
        db.session.commit()
        logging.info(f"Successfully created report field: id={new_field.id}, name={new_field.name}")
        return jsonify({
            'id': new_field.id,
            'name': new_field.name,
            'field_type': 'numeric',
            'formula': new_field.formula,
            'max_limit': new_field.max_limit
        }), 201
        
    except Exception as e:
        logging.error(f"Error managing report field: {str(e)}")
        logging.error(f"Request data: {data if 'data' in locals() else 'No data'}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'Failed to manage report field: {str(e)}'}), 500

@app.route("/task/<int:task_id>/units-completed", methods=['GET'])
def task_units_completed_route(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()

        if not company:
            return render_template('task_units_completed.html', task=None, attendance_records=[], workers=[], selected_date=None)

        # Get the specific task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        if not task or task.payment_type != 'per_part':
            return render_template('500.html'), 500

        # Get all available workers for the company for the dropdown
        available_workers = Worker.query.filter_by(company_id=company.id).all()

        # Get selected date from query parameter, default to today
        from datetime import date
        today = date.today()
        
        # Auto-update task status if start date has passed
        if task.start_date.date() <= today and task.status == 'Pending':
            task.status = 'In Progress'
            db.session.commit()
        
        selected_date_str = request.args.get('date')
        if selected_date_str:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        else:
            selected_date = date.today()

        # If selected date is before task start date, show error and do not show units UI
        if selected_date < task.start_date.date():
            return render_template('task_units_completed.html', 
                task=task, 
                attendance_records=[],
                workers=[],
                available_workers=available_workers,
                selected_date=selected_date,
                attendance_date_error="Units cannot be recorded before the task's start date."
            )

        # Get attendance records for the selected date
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        # Create attendance records for assigned workers if they don't exist
        existing_worker_ids = {record.worker_id for record in attendance_records}
        for worker in task.workers:
            if worker.id not in existing_worker_ids:
                new_attendance = Attendance(
                    worker_id=worker.id,
                    company_id=company.id,
                    date=selected_date,
                    status='Absent',
                    task_id=task.id
                )
                db.session.add(new_attendance)
        db.session.commit()

        # Fetch attendance records again to ensure up-to-date list
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        return render_template('task_units_completed.html', 
            task=task, 
            attendance_records=attendance_records,
            workers=task.workers.all(),  # Only show assigned workers
            available_workers=available_workers,  # Pass all available workers for dropdown
            selected_date=selected_date,
            attendance_date_error=None
        )
    except Exception as e:
        logging.error(f"Error fetching units completed: {str(e)}")
        return render_template('500.html'), 500

@app.route("/task/<int:task_id>/hours-worked", methods=['GET'])
def task_hours_worked_route(task_id):
    try:
        # Get current company from workspace
        company = get_current_company()

        if not company:
            return render_template('task_hours_worked.html', task=None, attendance_records=[], workers=[], selected_date=None)

        # Get the specific task
        task = Task.query.filter_by(id=task_id, company_id=company.id).first()
        if not task or task.payment_type != 'per_hour':
            return render_template('500.html'), 500

        # Get all available workers for the company for the dropdown
        available_workers = Worker.query.filter_by(company_id=company.id).all()

        # Get selected date from query parameter, default to today
        from datetime import date
        today = date.today()
        
        # Auto-update task status if start date has passed
        if task.start_date.date() <= today and task.status == 'Pending':
            task.status = 'In Progress'
            db.session.commit()
        
        selected_date_str = request.args.get('date')
        if selected_date_str:
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = today
        else:
            selected_date = today

        # Get attendance records for the selected date
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        # Ensure all assigned workers have attendance records for this date
        existing_worker_ids = {record.worker_id for record in attendance_records}
        for worker in task.workers:
            if worker.id not in existing_worker_ids:
                new_attendance = Attendance(
                    worker_id=worker.id,
                    company_id=company.id,
                    date=selected_date,
                    status='Absent',
                    task_id=task.id,
                    hours_worked=0
                )
                db.session.add(new_attendance)
        db.session.commit()

        # Fetch attendance records again to ensure up-to-date list
        attendance_records = Attendance.query.filter(
            Attendance.company_id == company.id,
            Attendance.date == selected_date,
            Attendance.task_id == task.id
        ).all()

        return render_template('task_hours_worked.html', 
            task=task, 
            attendance_records=attendance_records,
            workers=task.workers.all(),  # Only show assigned workers
            available_workers=available_workers,  # Pass all available workers for dropdown
            selected_date=selected_date,
            attendance_date_error=None
        )
    except Exception as e:
        logging.error(f"Error fetching hours worked: {str(e)}")
        return render_template('500.html'), 500

@app.route('/admin/master-dashboard')
@master_admin_required
def master_dashboard_route():
    """Master Admin Dashboard - Enhanced Platform Overview with Marketing Insights"""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import case, and_
        
        # Time periods for calculations
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        ninety_days_ago = datetime.utcnow() - timedelta(days=90)
        one_year_ago = datetime.utcnow() - timedelta(days=365)
        
        # 1. Enhanced Platform Summary Stats
        total_workspaces = Workspace.query.count()
        total_users = User.query.count()
        total_workers = Worker.query.count()
        total_tasks = Task.query.count()
        total_companies = Company.query.count()
        
        # Active metrics
        active_users_7d = User.query.filter(User.created_at >= seven_days_ago).count()
        active_users_30d = User.query.filter(User.created_at >= thirty_days_ago).count()
        recent_users = User.query.filter(User.created_at >= seven_days_ago).count()
        recent_tasks = Task.query.filter(Task.created_at >= seven_days_ago).count()
        recent_workers = Worker.query.filter(Worker.created_at >= seven_days_ago).count()
        
        # Task metrics this month
        current_month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        
        tasks_this_month = Task.query.filter(Task.created_at >= current_month_start).count()
        tasks_last_month = Task.query.filter(
            Task.created_at >= previous_month_start,
            Task.created_at < current_month_start
        ).count()
        
        # Calculate task growth percentage
        task_growth_percent = 0
        if tasks_last_month > 0:
            task_growth_percent = ((tasks_this_month - tasks_last_month) / tasks_last_month) * 100
        elif tasks_this_month > 0:
            task_growth_percent = 100
        
        # 2. Enhanced Workspace Analytics with Business Metrics
        # Get basic workspace data
        enhanced_workspaces_query = db.session.query(
            Workspace,
            func.count(UserWorkspace.id).label('user_count')
        ).outerjoin(UserWorkspace, Workspace.id == UserWorkspace.workspace_id)\
         .group_by(Workspace.id)\
         .order_by(desc(Workspace.created_at))\
         .all()
        
        # Convert to enhanced workspace data with additional metrics
        enhanced_workspaces = []
        most_active_workspace = None
        max_activity_score = 0
        
        for workspace, user_count in enhanced_workspaces_query:
            # Get company for this workspace
            company = Company.query.filter_by(workspace_id=workspace.id).first()
            
            if company:
                # Get worker count
                worker_count = Worker.query.filter_by(company_id=company.id).count()
                workers_this_month = Worker.query.filter(
                    Worker.company_id == company.id,
                    Worker.created_at >= current_month_start
                ).count()
                
                # Get task counts
                task_count = Task.query.filter_by(company_id=company.id).count()
                tasks_this_month = Task.query.filter(
                    Task.company_id == company.id,
                    Task.created_at >= current_month_start
                ).count()
                completed_tasks = Task.query.filter(
                    Task.company_id == company.id,
                    Task.status == 'Completed'
                ).count()
                
                # Get attendance data
                total_attendance_days = Attendance.query.filter(
                    Attendance.company_id == company.id,
                    Attendance.date >= current_month_start,
                    Attendance.status == 'Present'
                ).count()
            else:
                worker_count = 0
                workers_this_month = 0
                task_count = 0
                tasks_this_month = 0
                completed_tasks = 0
                total_attendance_days = 0
            
            activity_score = task_count + worker_count + tasks_this_month
            completion_rate = (completed_tasks / task_count * 100) if task_count > 0 else 0
            
            workspace_info = {
                'workspace': workspace,
                'user_count': user_count,
                'company_count': 1 if company else 0,
                'task_count': task_count,
                'worker_count': worker_count,
                'tasks_this_month': tasks_this_month,
                'workers_this_month': workers_this_month,
                'completed_tasks': completed_tasks,
                'total_attendance_days': total_attendance_days,
                'activity_score': activity_score,
                'completion_rate': completion_rate
            }
            enhanced_workspaces.append(workspace_info)
            
            # Track most active workspace
            if activity_score > max_activity_score:
                max_activity_score = activity_score
                most_active_workspace = workspace_info
        
        # Recent workspaces (top 10 for display)
        recent_workspaces = enhanced_workspaces[:10]
        
        # Industry and country stats
        industry_stats = db.session.query(
            Workspace.industry_type,
            func.count(Workspace.id).label('count')
        ).group_by(Workspace.industry_type).order_by(desc('count')).all()
        
        country_stats = db.session.query(
            Workspace.country,
            func.count(Workspace.id).label('count')
        ).group_by(Workspace.country).order_by(desc('count')).all()
        
        # Top performing workspaces by different metrics
        top_workspaces_by_tasks = sorted(enhanced_workspaces, key=lambda x: x['task_count'], reverse=True)[:5]
        top_workspaces_by_workers = sorted(enhanced_workspaces, key=lambda x: x['worker_count'], reverse=True)[:5]
        top_workspaces_by_activity = sorted(enhanced_workspaces, key=lambda x: x['activity_score'], reverse=True)[:5]
        
        # 3. Enhanced User Analytics
        role_stats = db.session.query(
            UserWorkspace.role,
            func.count(UserWorkspace.id).label('count')
        ).group_by(UserWorkspace.role).all()
        
        # User engagement metrics
        all_users_data = db.session.query(
            User,
            func.count(UserWorkspace.id).label('workspace_count')
        ).outerjoin(UserWorkspace, User.id == UserWorkspace.user_id)\
         .group_by(User.id)\
         .order_by(desc(User.id))\
         .limit(50).all()
        
        all_users = []
        for user_data in all_users_data:
            user, workspace_count = user_data
            # Get user roles
            user_roles = db.session.query(UserWorkspace.role).filter_by(user_id=user.id).distinct().all()
            roles = [role[0] for role in user_roles] if user_roles else []
            
            all_users.append({
                'user': user,
                'workspace_count': workspace_count,
                'roles': roles
            })
        
        # User growth data (last 30 days)
        user_growth_data = []
        task_growth_data = []
        for i in range(30):
            date = datetime.utcnow() - timedelta(days=i)
            user_count = User.query.filter(
                func.date(User.created_at) == date.date()
            ).count()
            task_count = Task.query.filter(
                func.date(Task.created_at) == date.date()
            ).count()
            user_growth_data.append((date.strftime('%m-%d'), user_count))
            task_growth_data.append((date.strftime('%m-%d'), task_count))
        user_growth_data.reverse()
        task_growth_data.reverse()
        
        # 4. Enhanced Subscription & Revenue Analytics
        subscription_stats = db.session.query(
            Workspace.subscription_status,
            func.count(Workspace.id).label('count')
        ).group_by(Workspace.subscription_status).all()
        
        active_subscriptions = sum(count for status, count in subscription_stats if status == 'active')
        free_trials = sum(count for status, count in subscription_stats if status == 'trial')
        failed_payments = sum(count for status, count in subscription_stats if status == 'past_due')
        
        # Enhanced revenue calculation
        estimated_revenue = active_subscriptions * 29.99
        
        # Expiring trials (next 7 days)
        expiring_trials = Workspace.query.filter(
            Workspace.trial_end_date <= datetime.utcnow() + timedelta(days=7),
            Workspace.subscription_status == 'trial'
        ).all()
        
        # 5. Marketing Analytics & Business Intelligence
        marketing_data = {
            'conversion_metrics': {
                'total_signups': total_workspaces,
                'paid_workspaces': active_subscriptions,
                'trial_workspaces': free_trials,
                'conversion_rate': round((active_subscriptions / total_workspaces * 100) if total_workspaces > 0 else 0, 1)
            },
            'geographic_data': [(country, count, 0, 0) for country, count in country_stats[:10]],
            'industry_performance': [],
            'growth_data': [],
            'market_opportunities': [],
            'industry_opportunities': []
        }
        
        # Industry performance analysis
        for industry, workspace_count in industry_stats:
            if workspace_count > 0:
                paid_in_industry = db.session.query(func.count(Workspace.id)).filter(
                    Workspace.industry_type == industry,
                    Workspace.subscription_status == 'active'
                ).scalar()
                
                marketing_data['industry_performance'].append((
                    industry, workspace_count, 0, paid_in_industry, 0
                ))
        
        # Growth trend data (last 6 months)
        for i in range(6):
            month_start = datetime.utcnow().replace(day=1) - timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            signups = Workspace.query.filter(
                Workspace.created_at >= month_start,
                Workspace.created_at < month_end
            ).count()
            
            conversions = Workspace.query.filter(
                Workspace.created_at >= month_start,
                Workspace.created_at < month_end,
                Workspace.subscription_status == 'active'
            ).count()
            
            marketing_data['growth_data'].append({
                'month': month_start.strftime('%Y-%m'),
                'signups': signups,
                'conversions': conversions,
                'conversion_rate': (conversions / signups * 100) if signups > 0 else 0
            })
        
        marketing_data['growth_data'].reverse()
        
        # Market opportunities (top countries with growth potential)
        for country, count in country_stats[:5]:
            potential_revenue = count * 29.99  # Assuming conversion
            marketing_data['market_opportunities'].append({
                'country': country,
                'workspace_count': count,
                'potential_revenue': potential_revenue
            })
        
        # Industry opportunities (best conversion rates)
        industry_conversions = []
        for industry, workspace_count in industry_stats:
            if workspace_count >= 3:  # Only consider industries with meaningful data
                paid_count = db.session.query(func.count(Workspace.id)).filter(
                    Workspace.industry_type == industry,
                    Workspace.subscription_status == 'active'
                ).scalar()
                
                conversion_rate = (paid_count / workspace_count * 100) if workspace_count > 0 else 0
                industry_conversions.append({
                    'industry': industry,
                    'workspace_count': workspace_count,
                    'conversion_rate': conversion_rate
                })
        
        marketing_data['industry_opportunities'] = sorted(
            industry_conversions, key=lambda x: x['conversion_rate'], reverse=True
        )[:3]
        
        marketing_data['active_users_30d'] = active_users_30d
        
        # 6. Business Health Metrics
        inactive_workspaces = sum(1 for ws in enhanced_workspaces if ws['tasks_this_month'] == 0)
        high_activity_workspaces = sum(1 for ws in enhanced_workspaces if ws['activity_score'] > 10)
        
        # 7. Enhanced Alerts & Notifications
        alerts = []
        
        if expiring_trials:
            alerts.append({
                'type': 'warning',
                'message': f'{len(expiring_trials)} workspaces have expiring trials in the next 7 days'
            })
        
        if failed_payments > 0:
            alerts.append({
                'type': 'error',
                'message': f'{failed_payments} workspaces have failed payments'
            })
        
        if inactive_workspaces > 0:
            alerts.append({
                'type': 'info',
                'message': f'{inactive_workspaces} workspaces have no tasks created this month'
            })
        
        # 8. Top Performing Metrics for Marketing
        top_workspaces = sorted(enhanced_workspaces, key=lambda x: x['task_count'], reverse=True)[:5]
        
        # Additional master admin stats
        master_admins_count = MasterAdmin.query.filter_by(is_active=True).count()
        
        # 9. Enhanced Task Analytics with Worker Counts
        # Get tasks with most workers (through Task-Worker relationships)
        task_worker_analytics = []
        
        # Query tasks with worker counts from attendance records
        task_worker_counts = db.session.query(
            Task.id,
            Task.name,
            Task.payment_type,
            Task.status,
            Task.created_at,
            Company.name.label('company_name'),
            Workspace.name.label('workspace_name'),
            func.count(func.distinct(Attendance.worker_id)).label('unique_workers'),
            func.count(Attendance.id).label('total_attendance'),
            func.avg(case(
                (Attendance.units_completed.isnot(None), Attendance.units_completed),
                else_=0
            )).label('avg_units_completed')
        ).join(Company, Task.company_id == Company.id)\
         .join(Workspace, Company.workspace_id == Workspace.id)\
         .outerjoin(Attendance, Task.id == Attendance.task_id)\
         .group_by(Task.id, Task.name, Task.payment_type, Task.status, Task.created_at, Company.name, Workspace.name)\
         .order_by(desc('unique_workers'), desc('total_attendance'))\
         .limit(20).all()
        
        for task_data in task_worker_counts:
            task_id, task_name, payment_type, status, created_at, company_name, workspace_name, unique_workers, total_attendance, avg_units = task_data
            
            # Calculate task priority based on worker engagement and completion
            if unique_workers >= 10:
                priority = 'high'
            elif unique_workers >= 5:
                priority = 'medium'
            else:
                priority = 'low'
            
            # Calculate task efficiency metrics
            efficiency_score = 0
            if total_attendance > 0 and avg_units:
                efficiency_score = round(float(avg_units) * unique_workers / total_attendance * 100, 1)
            
            task_worker_analytics.append({
                'task_id': task_id,
                'task_name': task_name,
                'payment_type': payment_type,
                'status': status,
                'created_at': created_at,
                'company_name': company_name,
                'workspace_name': workspace_name,
                'unique_workers': unique_workers or 0,
                'total_attendance': total_attendance or 0,
                'avg_units_completed': round(float(avg_units), 2) if avg_units else 0,
                'priority': priority,
                'efficiency_score': efficiency_score
            })
        
        # Top tasks by different metrics
        top_tasks_by_workers = sorted(task_worker_analytics, key=lambda x: x['unique_workers'], reverse=True)[:10]
        top_tasks_by_attendance = sorted(task_worker_analytics, key=lambda x: x['total_attendance'], reverse=True)[:10]
        top_tasks_by_efficiency = sorted(task_worker_analytics, key=lambda x: x['efficiency_score'], reverse=True)[:10]
        
        # Task completion analytics
        completed_tasks_with_workers = [t for t in task_worker_analytics if t['status'] == 'Completed']
        active_tasks_with_workers = [t for t in task_worker_analytics if t['status'] in ['Active', 'In Progress']]
        
        # Payment type distribution for high-worker tasks
        payment_type_analytics = {}
        for task in task_worker_analytics:
            ptype = task['payment_type'] or 'Not Set'
            if ptype not in payment_type_analytics:
                payment_type_analytics[ptype] = {
                    'count': 0,
                    'total_workers': 0,
                    'avg_workers': 0
                }
            payment_type_analytics[ptype]['count'] += 1
            payment_type_analytics[ptype]['total_workers'] += task['unique_workers']
        
        # Calculate averages
        for ptype in payment_type_analytics:
            if payment_type_analytics[ptype]['count'] > 0:
                payment_type_analytics[ptype]['avg_workers'] = round(
                    payment_type_analytics[ptype]['total_workers'] / payment_type_analytics[ptype]['count'], 1
                )
        
        # Monthly task trends with worker engagement
        monthly_task_trends = []
        for i in range(12):
            month_start = datetime.utcnow().replace(day=1) - timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            tasks_created = Task.query.filter(
                Task.created_at >= month_start,
                Task.created_at < month_end
            ).count()
            
            # Get worker engagement for tasks created in this month
            worker_engagement = db.session.query(
                func.count(func.distinct(Attendance.worker_id))
            ).join(Task, Attendance.task_id == Task.id)\
             .filter(
                Task.created_at >= month_start,
                Task.created_at < month_end
            ).scalar() or 0
            
            monthly_task_trends.append({
                'month': month_start.strftime('%Y-%m'),
                'tasks_created': tasks_created,
                'worker_engagement': worker_engagement,
                'avg_workers_per_task': round(worker_engagement / tasks_created, 1) if tasks_created > 0 else 0
            })
        
        monthly_task_trends.reverse()
        
        return render_template('admin/master_dashboard.html',
                             # Enhanced Platform Summary
                             total_workspaces=total_workspaces,
                             total_users=total_users,
                             total_companies=total_companies,
                             active_users_7d=active_users_7d,
                             active_users_30d=active_users_30d,
                             total_workers=total_workers,
                             total_tasks=total_tasks,
                             tasks_this_month=tasks_this_month,
                             task_growth_percent=task_growth_percent,
                             recent_users=recent_users,
                             recent_tasks=recent_tasks,
                             recent_workers=recent_workers,
                             
                             # Enhanced Workspace Analytics
                             industry_stats=industry_stats,
                             country_stats=country_stats,
                             recent_workspaces=recent_workspaces,
                             enhanced_workspaces=enhanced_workspaces[:20],  # Limit for display
                             most_active_workspace=most_active_workspace,
                             top_workspaces_by_tasks=top_workspaces_by_tasks,
                             top_workspaces_by_workers=top_workspaces_by_workers,
                             top_workspaces_by_activity=top_workspaces_by_activity,
                             inactive_workspaces=inactive_workspaces,
                             high_activity_workspaces=high_activity_workspaces,
                             
                             # Enhanced User Analytics
                             role_stats=role_stats,
                             user_growth_data=user_growth_data,
                             task_growth_data=task_growth_data,
                             all_users=all_users,
                             
                             # Enhanced Subscription & Revenue
                             subscription_stats=subscription_stats,
                             active_subscriptions=active_subscriptions,
                             free_trials=free_trials,
                             failed_payments=failed_payments,
                             estimated_revenue=estimated_revenue,
                             expiring_trials=expiring_trials,
                             
                             # Marketing Analytics
                             marketing_data=marketing_data,
                             
                             # Top Performing Metrics
                             top_workspaces=top_workspaces,
                             
                             # Management
                             master_admins_count=master_admins_count,
                             
                             # Enhanced Task Analytics
                             task_worker_analytics=task_worker_analytics,
                             top_tasks_by_workers=top_tasks_by_workers,
                             top_tasks_by_attendance=top_tasks_by_attendance,
                             top_tasks_by_efficiency=top_tasks_by_efficiency,
                             completed_tasks_with_workers=completed_tasks_with_workers,
                             active_tasks_with_workers=active_tasks_with_workers,
                             payment_type_analytics=payment_type_analytics,
                             monthly_task_trends=monthly_task_trends,
                             
                             # Alerts
                             alerts=alerts,
                             
                             # Current time for calculations
                             now=datetime.utcnow())
    except Exception as e:
        logging.error(f"Error in master dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        return render_template('500.html'), 500

@app.route('/admin/user-growth-data/<period>')
@master_admin_required
def user_growth_data_route(period):
    try:
        # Determine the number of days and format based on period
        if period == '7d':
            days = 7
            date_format = '%m-%d'
        elif period == '30d':
            days = 30
            date_format = '%m-%d'
        elif period == '1y':
            days = 365
            date_format = '%Y-%m'  # Group by month for yearly view
        else:
            return jsonify({'error': 'Invalid period'}), 400
        
        user_growth_data = []
        
        if period == '1y':
            # For yearly view, group by month
            for i in range(12):
                start_date = datetime.utcnow().replace(day=1) - timedelta(days=30 * i)
                end_date = start_date.replace(day=1) + timedelta(days=32)
                end_date = end_date.replace(day=1) - timedelta(days=1)  # Last day of month
                
                user_count = User.query.filter(
                    func.date(User.created_at) >= start_date.date(),
                    func.date(User.created_at) <= end_date.date()
                ).count()
                
                user_growth_data.append({
                    'label': start_date.strftime('%Y-%m'),
                    'count': user_count
                })
            user_growth_data.reverse()
        else:
            # For daily view (7d and 30d)
            for i in range(days):
                date = datetime.utcnow() - timedelta(days=i)
                user_count = User.query.filter(
                    func.date(User.created_at) == date.date()
                ).count()
                user_growth_data.append({
                    'label': date.strftime(date_format),
                    'count': user_count
                })
            user_growth_data.reverse()
        
        return jsonify({
            'success': True,
            'data': user_growth_data
        })
        
    except Exception as e:
        logging.error(f"Error fetching user growth data: {str(e)}")
        return jsonify({'error': f'Error fetching data: {str(e)}'}), 500

@app.route('/admin/pause-workspace/<int:workspace_id>', methods=['POST'])
@master_admin_required
def pause_workspace_route(workspace_id):
    try:
        workspace = Workspace.query.get(workspace_id)
        if not workspace:
            return jsonify({'error': 'Workspace not found'}), 404
        
        # Store original subscription status and set to paused
        if workspace.subscription_status != 'paused':
            # We'll use 'paused' as a special status
            workspace.subscription_status = 'paused'
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': f'Workspace "{workspace.name}" has been paused',
                'status': 'paused'
            })
        else:
            return jsonify({'error': 'Workspace is already paused'}), 400
            
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error pausing workspace: {str(e)}")
        return jsonify({'error': f'Error pausing workspace: {str(e)}'}), 500

@app.route('/admin/resume-workspace/<int:workspace_id>', methods=['POST'])
@master_admin_required
def resume_workspace_route(workspace_id):
    try:
        workspace = Workspace.query.get(workspace_id)
        if not workspace:
            return jsonify({'error': 'Workspace not found'}), 404
        
        if workspace.subscription_status == 'paused':
            # Resume with appropriate status based on subscription dates
            from datetime import datetime
            now = datetime.utcnow()
            
            if workspace.subscription_end_date and workspace.subscription_end_date > now:
                workspace.subscription_status = 'active'
            elif workspace.trial_end_date and workspace.trial_end_date > now:
                workspace.subscription_status = 'trial'
            else:
                workspace.subscription_status = 'expired'
                
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': f'Workspace "{workspace.name}" has been resumed',
                'status': workspace.subscription_status
            })
        else:
            return jsonify({'error': 'Workspace is not paused'}), 400
            
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error resuming workspace: {str(e)}")
        return jsonify({'error': f'Error resuming workspace: {str(e)}'}), 500

@app.route('/admin/delete-workspace/<int:workspace_id>', methods=['POST'])
@master_admin_required
def delete_workspace_route(workspace_id):
    try:
        workspace = Workspace.query.get(workspace_id)
        if not workspace:
            return jsonify({'error': 'Workspace not found'}), 404
        
        # With proper cascade relationships, we can simply delete the workspace
        # and all related data will be automatically deleted
        db.session.delete(workspace)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error deleting workspace: {str(e)}")
        return jsonify({'error': f'Error deleting workspace: {str(e)}'}), 500

@app.route('/admin/delete-worker/<int:worker_id>', methods=['POST'])
@master_admin_required
def delete_worker_route(worker_id):
    try:
        worker = Worker.query.get(worker_id)
        if not worker:
            return jsonify({'error': 'Worker not found'}), 404
        db.session.delete(worker)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/master-admins')
@master_admin_required
def master_admins_route():
    """Master Admins Management Page"""
    try:
        master_admins = MasterAdmin.query.filter_by(is_active=True).all()
        return render_template('admin/master_admins.html', master_admins=master_admins)
    except Exception as e:
        logging.error(f"Error in master admins route: {str(e)}")
        return render_template('500.html'), 500

@app.route('/admin/add-master-admin', methods=['POST'])
@master_admin_required
def add_master_admin_route():
    """Add a new master admin"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        name = data.get('name', '').strip()
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Check if email already exists
        existing_admin = MasterAdmin.query.filter_by(email=email).first()
        if existing_admin:
            return jsonify({'error': 'Email already exists'}), 400
        
        # Get current master admin
        current_admin = MasterAdmin.query.filter_by(email=session['user']['user_email']).first()
        
        new_master_admin = MasterAdmin(
            email=email,
            name=name,
            created_by=current_admin.id if current_admin else None
        )
        
        db.session.add(new_master_admin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Master admin added successfully'
        })
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error adding master admin: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/delete-master-admin/<int:admin_id>', methods=['POST'])
@master_admin_required
def delete_master_admin_route(admin_id):
    """Delete a master admin (soft delete)"""
    try:
        master_admin = MasterAdmin.query.get(admin_id)
        if not master_admin:
            return jsonify({'error': 'Master admin not found'}), 404
        
        # Don't allow self-deletion
        if master_admin.email == session['user']['user_email']:
            return jsonify({'error': 'Cannot delete yourself'}), 400
        
        master_admin.is_active = False
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@master_admin_required
def delete_user_route(user_id):
    """Delete a user"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export/workspaces')
@master_admin_required
def export_workspaces_route():
    """Export workspaces to CSV"""
    try:
        import csv
        from io import StringIO
        from datetime import datetime
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['ID', 'Name', 'Country', 'Industry', 'Workspace Code', 'Subscription Status', 'Created At', 'User Count'])
        
        # Get all workspaces with user counts
        workspaces = db.session.query(
            Workspace,
            func.count(UserWorkspace.id).label('user_count')
        ).outerjoin(UserWorkspace, Workspace.id == UserWorkspace.workspace_id)\
         .group_by(Workspace.id)\
         .all()
        
        for workspace, user_count in workspaces:
            writer.writerow([
                workspace.id,
                workspace.name,
                workspace.country,
                workspace.industry_type,
                workspace.workspace_code,
                workspace.subscription_status,
                workspace.created_at.strftime('%Y-%m-%d'),
                user_count
            ])
        
        output.seek(0)
        return send_file(
            StringIO(output.getvalue()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'workspaces_export_{datetime.utcnow().strftime("%Y%m%d")}.csv'
        )
    except Exception as e:
        logging.error(f"Error exporting workspaces: {str(e)}")
        return jsonify({'error': 'Failed to export workspaces'}), 500

@app.route('/admin/export/users')
@master_admin_required
def export_users_route():
    """Export users to CSV"""
    try:
        import csv
        from io import StringIO
        from datetime import datetime
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['ID', 'Email', 'Profile Picture', 'Role', 'Created At'])
        
        users = User.query.all()
        for user in users:
            writer.writerow([
                user.id,
                user.email,
                user.profile_picture or '',
                user.role,
                user.created_at.strftime('%Y-%m-%d') if hasattr(user, 'created_at') else ''
            ])
        
        output.seek(0)
        return send_file(
            StringIO(output.getvalue()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'users_export_{datetime.utcnow().strftime("%Y%m%d")}.csv'
        )
    except Exception as e:
        logging.error(f"Error exporting users: {str(e)}")
        return jsonify({'error': 'Failed to export users'}), 500

@app.route('/admin/export/revenue')
@master_admin_required
def export_revenue_route():
    """Export revenue report to CSV"""
    try:
        import csv
        from io import StringIO
        from datetime import datetime
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Workspace', 'Country', 'Industry', 'Subscription Status', 'Trial End Date', 'Estimated Revenue'])
        
        workspaces = Workspace.query.all()
        for workspace in workspaces:
            revenue = 29.99 if workspace.subscription_status == 'active' else 0
            writer.writerow([
                workspace.name,
                workspace.country,
                workspace.industry_type,
                workspace.subscription_status,
                workspace.trial_end_date.strftime('%Y-%m-%d') if workspace.trial_end_date else '',
                f"${revenue:.2f}"
            ])
        
        output.seek(0)
        return send_file(
            StringIO(output.getvalue()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'revenue_report_{datetime.utcnow().strftime("%Y%m%d")}.csv'
        )
    except Exception as e:
        logging.error(f"Error exporting revenue: {str(e)}")
        return jsonify({'error': 'Failed to export revenue report'}), 500

@app.route('/admin/debug-master-dashboard')
def debug_master_dashboard():
    """Debug route to check master dashboard data and authentication"""
    try:
        from datetime import datetime, timedelta
        
        # Check authentication
        auth_status = {
            'session_exists': 'user' in session,
            'user_email': session.get('user', {}).get('user_email') if 'user' in session else None,
            'is_master_admin': False
        }
        
        # Check if user is master admin
        if 'user' in session and 'user_email' in session['user']:
            from models import MasterAdmin
            master_admin = MasterAdmin.query.filter_by(
                email=session['user']['user_email'], 
                is_active=True
            ).first()
            auth_status['is_master_admin'] = bool(master_admin)
            auth_status['master_admin_record'] = master_admin.to_dict() if master_admin else None
        
        # Get basic counts
        counts = {
            'total_workspaces': Workspace.query.count(),
            'total_users': User.query.count(),
            'total_workers': Worker.query.count(),
            'total_tasks': Task.query.count(),
            'total_master_admins': MasterAdmin.query.count()
        }
        
        # Get sample data
        sample_data = {
            'recent_workspaces': [w.to_dict() for w in Workspace.query.order_by(Workspace.created_at.desc()).limit(5).all()],
            'recent_users': [u.to_dict() for u in User.query.order_by(User.created_at.desc()).limit(5).all()],
            'master_admins': [ma.to_dict() for ma in MasterAdmin.query.all()]
        }
        
        return jsonify({
            'auth_status': auth_status,
            'counts': counts,
            'sample_data': sample_data,
            'timestamp': datetime.utcnow().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/admin/create-master-admin-debug', methods=['POST'])
def create_master_admin_debug():
    """Debug route to create a master admin"""
    try:
        from models import MasterAdmin
        
        # Get email from request
        data = request.get_json()
        email = data.get('email', 'markbmwape@gmail.com')
        name = data.get('name', 'Master Admin')
        
        # Check if master admin already exists
        existing = MasterAdmin.query.filter_by(email=email).first()
        if existing:
            return jsonify({
                'message': 'Master admin already exists',
                'master_admin': existing.to_dict()
            })
        
        # Create new master admin
        master_admin = MasterAdmin(
            email=email,
            name=name,
            is_active=True
        )
        db.session.add(master_admin)
        db.session.commit()
        
        return jsonify({
            'message': 'Master admin created successfully',
            'master_admin': master_admin.to_dict()
        })
    except Exception as e:
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/firebase_config')
def get_firebase_config():
    try:
        # Load Firebase configuration from environment variable or use default config
        firebase_config_str = os.environ.get('FIREBASE_CONFIG', '')
        
        if firebase_config_str:
            firebase_config = json.loads(firebase_config_str)
        else:
            # Get the current domain for authDomain
            current_domain = request.host
            auth_domain = "ember-accounting.firebaseapp.com"
            
            # For local development, use localhost
            if 'localhost' in current_domain or '127.0.0.1' in current_domain:
                auth_domain = "ember-accounting.firebaseapp.com"
            # For production, you might want to use your custom domain
            elif 'run.app' in current_domain:
                auth_domain = "ember-accounting.firebaseapp.com"  # Keep Firebase domain for now
            
            # New Firebase configuration for ember-accounting project
            firebase_config = {
                "apiKey": "AIzaSyD9_N_0ve9ABFwwnTBn1N2oxlUs6xbT-No",
                "authDomain": auth_domain,
                "projectId": "ember-accounting",
                "storageBucket": "ember-accounting.firebasestorage.app",
                "messagingSenderId": "328324461979",
                "appId": "1:328324461979:web:0cc9ddad6aa3f157359d3e",
                "measurementId": "G-F1XTE0TP63"
            }
        
        logging.info(f"Firebase config authDomain: {firebase_config.get('authDomain')}")
        return jsonify(firebase_config)
    except Exception as e:
        logging.error(f"Error retrieving Firebase config: {str(e)}")
        # Return the new fallback configuration
        return jsonify({
            "apiKey": "AIzaSyD9_N_0ve9ABFwwnTBn1N2oxlUs6xbT-No",
            "authDomain": "ember-accounting.firebaseapp.com",
            "projectId": "ember-accounting",
            "storageBucket": "ember-accounting.firebasestorage.app",
            "messagingSenderId": "328324461979",
            "appId": "1:328324461979:web:0cc9ddad6aa3f157359d3e",
            "measurementId": "G-F1XTE0TP63"
        })